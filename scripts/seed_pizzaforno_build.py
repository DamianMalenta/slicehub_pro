#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/seed_pizzaforno_build.py
================================
Czyta menu (14).xlsx + additions.xlsx z _docs/menu_pizzaforno/
i generuje scripts/seed_pizzaforno.sql — produkcyjny seed Pizzerii Forno.

Uruchomienie:
    pip install openpyxl
    python3 scripts/seed_pizzaforno_build.py
    # → scripts/seed_pizzaforno.sql

Domyślnie tenant_id=2 (Pizza Forno na slicehub.net).
"""

import openpyxl
import re
import uuid
import datetime
import random
import sys
import os
from decimal import Decimal, ROUND_HALF_UP

# ──────────────────────────────────────────────────────────────────────────────
# PATHS
# ──────────────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MENU_FILE    = os.path.join(BASE_DIR, '_docs', 'menu_pizzaforno', 'menu (14).xlsx')
ADD_FILE     = os.path.join(BASE_DIR, '_docs', 'menu_pizzaforno', 'additions.xlsx')
OUTPUT_FILE  = os.path.join(BASE_DIR, 'scripts', 'seed_pizzaforno.sql')
MENU_OUTPUT_FILE = os.path.join(BASE_DIR, 'scripts', 'seed_pizzaforno_menu.sql')
OPS_OUTPUT_FILE  = os.path.join(BASE_DIR, 'scripts', 'seed_pizzaforno_ops.sql')
DEFAULT_TID  = 2
WAREHOUSE_ID = 'MAIN'

# Precomputed bcrypt hash for password 'password' (used by ops-only seed users)
BCRYPT_HASH = '$2y$10$92IXUNpkjO0rOQ5byMi.Ye4oKoEa3Ro9llC/.og/at2.uheWG/igi'

random.seed(42)  # reproducible

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

DIACRITICS = {
    'ą': 'A', 'ć': 'C', 'ę': 'E', 'ł': 'L', 'ń': 'N',
    'ó': 'O', 'ś': 'S', 'ź': 'Z', 'ż': 'Z',
    'Ą': 'A', 'Ć': 'C', 'Ę': 'E', 'Ł': 'L', 'Ń': 'N',
    'Ó': 'O', 'Ś': 'S', 'Ź': 'Z', 'Ż': 'Z',
}


def make_sku(name: str) -> str:
    """Convert Polish name → ASCII uppercase SKU key."""
    result = name
    for k, v in DIACRITICS.items():
        result = result.replace(k, v)
    result = result.upper()
    result = re.sub(r'[^A-Z0-9]+', '_', result)
    result = re.sub(r'_+', '_', result)
    result = result.strip('_')
    return result


def esc(s) -> str:
    """Escape string for SQL single-quote context."""
    if s is None:
        return 'NULL'
    s = str(s)
    s = s.replace('\\', '\\\\').replace("'", "\\'")
    return f"'{s}'"


def esc_float(v, digits=2) -> str:
    if v is None:
        return 'NULL'
    return str(round(float(v), digits))


def gen_uuid() -> str:
    return str(uuid.uuid4())


def days_ago(n: int) -> str:
    dt = datetime.datetime.now() - datetime.timedelta(days=n)
    return dt.strftime('%Y-%m-%d %H:%M:%S')


def date_ago(n: int) -> str:
    dt = datetime.date.today() - datetime.timedelta(days=n)
    return dt.strftime('%Y-%m-%d')


def hours_ago(h: float) -> str:
    dt = datetime.datetime.now() - datetime.timedelta(hours=h)
    return dt.strftime('%Y-%m-%d %H:%M:%S')


def sql_hours_ago(h: float) -> str:
    """SQL expression — relative timestamp (re-seed stays fresh)."""
    total_minutes = max(1, int(round(h * 60)))
    if total_minutes % 60 == 0 and total_minutes >= 60:
        return f"DATE_SUB(NOW(), INTERVAL {total_minutes // 60} HOUR)"
    return f"DATE_SUB(NOW(), INTERVAL {total_minutes} MINUTE)"


def canonical_order_type(raw: str) -> str:
    return {'collection': 'takeaway', 'table': 'dine_in'}.get(raw, raw)


def canonical_payment_status(payment_status: str, payment_method: str | None) -> str:
    if payment_status in ('to_pay', 'online_unpaid', 'cash', 'card', 'online_paid'):
        return payment_status
    if payment_status == 'unpaid':
        return 'online_unpaid' if payment_method == 'online' else 'to_pay'
    if payment_status == 'paid':
        mapping = {'cash': 'cash', 'card': 'card', 'online': 'online_paid'}
        return mapping.get(payment_method or '', 'cash')
    return payment_status


def canonical_status_pair(status: str, order_type: str, delivery_status: str | None = None) -> tuple[str, str | None]:
    """Map legacy seed statuses to 3-pillar model (status + delivery_status)."""
    if delivery_status is not None:
        return status, delivery_status
    ot = canonical_order_type(order_type)
    if status == 'delivered':
        return 'completed', 'delivered'
    if status in ('in_route', 'in_delivery'):
        return 'ready', 'in_delivery'
    if ot != 'delivery':
        return status, None
    if status == 'completed':
        return status, 'delivered'
    if status == 'cancelled':
        return status, None
    return status, 'unassigned'


# ──────────────────────────────────────────────────────────────────────────────
# XLSX PARSERS
# ──────────────────────────────────────────────────────────────────────────────

def parse_menu_xlsx():
    """Read menu (14).xlsx → list of dicts (deduplicated by SKU, canonical category preferred)."""
    wb = openpyxl.load_workbook(MENU_FILE)
    ws = wb.active
    raw = []
    for r in range(2, ws.max_row + 1):
        name     = ws.cell(r, 2).value
        category = ws.cell(r, 4).value
        price    = ws.cell(r, 5).value
        vat      = ws.cell(r, 6).value
        desc     = ws.cell(r, 7).value
        weight   = ws.cell(r, 8).value
        prep     = ws.cell(r, 9).value
        if not name or not category:
            continue
        try:
            price_f = float(price) if price is not None else 0.0
        except (ValueError, TypeError):
            print(f"  [WARN] Cannot parse price for {name!r}: {price!r}", file=sys.stderr)
            price_f = 0.0
        raw.append({
            'name':     name.strip(),
            'category': category.strip(),
            'price':    price_f,
            'vat':      str(vat).strip() if vat else '',
            'desc':     desc.strip() if desc else '',
            'weight':   weight,
            'prep':     prep,
        })

    # Deduplicate by SKU — prefer the canonical category over NOWOŚCI !!!
    # Rule: if an item appears in multiple categories, the non-NOWOŚCI version wins.
    # For pizza variants (PIZZA 30/37 CM), dedup by base name + size.
    seen_skus = {}  # sku → item dict index in `items`
    items = []
    for item in raw:
        cat  = item['category']
        name = item['name']
        # Compute effective SKU (same logic as make_sku, stripping size suffixes)
        sku = make_sku(name)
        if sku not in seen_skus:
            seen_skus[sku] = len(items)
            items.append(item)
        else:
            # Duplicate — keep canonical (non-NOWOŚCI) version
            existing_idx = seen_skus[sku]
            existing = items[existing_idx]
            if existing['category'] == 'NOWOŚCI !!!' and cat != 'NOWOŚCI !!!':
                # Replace NOWOŚCI entry with canonical category entry
                items[existing_idx] = item
                print(f"  [DEDUP] {name!r}: keeping {cat!r} over 'NOWOŚCI !!!'", file=sys.stderr)
            else:
                print(f"  [DEDUP] {name!r}: skipping duplicate in {cat!r} (kept {existing['category']!r})", file=sys.stderr)

    return items


def parse_additions_xlsx():
    """Read additions.xlsx → list of modifier dicts."""
    wb = openpyxl.load_workbook(ADD_FILE)
    ws = wb.active
    mods = []
    for r in range(2, ws.max_row + 1):
        name     = ws.cell(r, 2).value
        price    = ws.cell(r, 3).value
        pos      = ws.cell(r, 4).value
        set_name = ws.cell(r, 6).value
        active   = ws.cell(r, 8).value
        typ      = ws.cell(r, 9).value
        required = ws.cell(r, 10).value
        countable= ws.cell(r, 11).value
        mn       = ws.cell(r, 12).value
        mx       = ws.cell(r, 13).value
        if not name or not set_name:
            continue
        try:
            price_f = float(price) if price is not None else 0.0
        except (ValueError, TypeError):
            print(f"  [WARN] Cannot parse modifier price for {name!r}: {price!r}", file=sys.stderr)
            price_f = 0.0
        mods.append({
            'name':     name.strip(),
            'price':    price_f,
            'position': int(pos) if pos is not None else 0,
            'set_name': set_name.strip(),
            'active':   str(active).strip().upper() == 'Y',
            'type':     str(typ).strip() if typ else 'multiple',
            'required': str(required).strip().upper() == 'Y',
            'countable':str(countable).strip().upper() == 'Y',
            'min':      int(mn) if mn else 0,
            'max':      int(mx) if mx else 0,
        })
    return mods


# ──────────────────────────────────────────────────────────────────────────────
# VARIANT FAMILIES
# ──────────────────────────────────────────────────────────────────────────────

def extract_variant_families(items):
    """
    Groups items into variant families:
      - PIZZA 30CM / 37CM → scale 'pizza_round'
      - PANINI MAŁE / DUŻE → scale 'panini_size'

    Returns:
        pizza_families: dict {base_name: {'30': item_dict, '37': item_dict}}
        panini_families: dict {base_name: {'MALE': item_dict, 'DUZE': item_dict}}
        singles: list of non-variant items (with possibly fixed category)
    """
    pizza_30 = {}
    pizza_37 = {}
    panini_male = {}
    panini_duze = {}
    singles = []

    for item in items:
        cat  = item['category']
        name = item['name']

        if cat == 'PIZZA 30 CM':
            base = re.sub(r'\s*30\s*cm\s*$', '', name, flags=re.IGNORECASE).strip()
            pizza_30[base] = item
        elif cat == 'PIZZA 37 CM':
            base = re.sub(r'\s*37\s*cm\s*$', '', name, flags=re.IGNORECASE).strip()
            pizza_37[base] = item
        elif cat == 'PANINI' and re.search(r'\s*-\s*MAŁE\s*$', name):
            base = re.sub(r'\s*-\s*MAŁE\s*$', '', name).strip()
            panini_male[base] = item
        elif cat == 'PANINI' and re.search(r'\s*-\s*DUŻE\s*$', name):
            base = re.sub(r'\s*-\s*DUŻE\s*$', '', name).strip()
            panini_duze[base] = item
        else:
            singles.append(item)

    pizza_families = {}
    all_pizza_bases = sorted(set(list(pizza_30.keys()) + list(pizza_37.keys())))
    for base in all_pizza_bases:
        pizza_families[base] = {
            '30': pizza_30.get(base),
            '37': pizza_37.get(base),
        }

    panini_families = {}
    all_panini_bases = sorted(set(list(panini_male.keys()) + list(panini_duze.keys())))
    for base in all_panini_bases:
        panini_families[base] = {
            'MALE': panini_male.get(base),
            'DUZE': panini_duze.get(base),
        }

    return pizza_families, panini_families, singles


# ──────────────────────────────────────────────────────────────────────────────
# SYS_ITEMS — ingredient dictionary
# ──────────────────────────────────────────────────────────────────────────────

# Master ingredient dictionary — ~70 SKU
# Format: (sku, name, base_unit, category_hint)
SYS_ITEMS = [
    # MĄKI / CISTA
    ('MAKA_TYP_00',       'Mąka pszenna typ 00',         'kg', 'suche'),
    ('MAKA_SEZAMOWA',     'Mąka sezamowa (ciasto)',       'kg', 'suche'),

    # SOSY BAZOWE
    ('SOS_POMIDOROWY',    'Sos pomidorowy passata',       'kg', 'sosy'),
    ('SOS_SMIETANKOWY',   'Sos śmietankowy',              'kg', 'sosy'),
    ('SOS_CZOSNKOWY',     'Sos czosnkowy',                'kg', 'sosy'),
    ('SOS_BBQ',           'Sos BBQ',                      'kg', 'sosy'),
    ('SOS_MEKSYKANSKI',   'Sos meksykański ostry',        'kg', 'sosy'),
    ('SOS_1000_WYSP',     'Sos 1000 wysp',                'kg', 'sosy'),
    ('TABASCO',           'Tabasco',                      'l',  'sosy'),
    ('MAJONEZ',           'Majonez',                      'kg', 'sosy'),
    ('KETCHUP',           'Ketchup',                      'kg', 'sosy'),
    ('KREM_BALSAMICZNY',  'Krem balsamiczny',             'l',  'sosy'),

    # SERY
    ('MOZZ_FIOR',         'Mozzarella fior di latte',     'kg', 'sery'),
    ('MOZZ_BUFFALO',      'Mozzarella di buffalo',        'kg', 'sery'),
    ('RICOTTA',           'Ricotta',                      'kg', 'sery'),
    ('PARMEZAN',          'Parmezan',                     'kg', 'sery'),
    ('GORGONZOLA',        'Gorgonzola',                   'kg', 'sery'),
    ('FETA',              'Feta',                         'kg', 'sery'),
    ('EDAMSKI',           'Ser Edamski',                  'kg', 'sery'),

    # MIĘSA
    ('SALAMI_PICANTE',    'Salami picante',               'kg', 'mięsa'),
    ('SALAMI',            'Salami',                       'kg', 'mięsa'),
    ('NDUJA',             'Nduja (włoska kiełbasa)',       'kg', 'mięsa'),
    ('KIELB_WLOSKA',      'Włoska kiełbasa',              'kg', 'mięsa'),
    ('SZYNKA_PARM',       'Szynka parmeńska',             'kg', 'mięsa'),
    ('SZYNKA',            'Szynka',                       'kg', 'mięsa'),
    ('KURCZAK',           'Kurczak (pierś)',               'kg', 'mięsa'),
    ('STEK_WOLOWY',       'Stek wołowy',                  'kg', 'mięsa'),
    ('BOCZEK',            'Boczek',                       'kg', 'mięsa'),
    ('RWANA_WIEPRZ',      'Rwana wieprzowina',            'kg', 'mięsa'),
    ('KEBAB_DROBIOWY',    'Drobiowy kebab',               'kg', 'mięsa'),
    ('MIESO_WOLOWE',      'Mięso wołowe',                 'kg', 'mięsa'),
    ('ANCHOIS',           'Fileciki anchois',             'kg', 'mięsa'),
    ('TUNCZYK',           'Tuńczyk (puszka)',             'kg', 'mięsa'),
    ('GYROS_MIESO',       'Gyros (mięso mieszane)',       'kg', 'mięsa'),

    # WARZYWA
    ('PIECZARKI',         'Pieczarki',                    'kg', 'warzywa'),
    ('PIECZARKI_SMAZONE', 'Smażone pieczarki',            'kg', 'warzywa'),
    ('CEBULA_CZERW',      'Czerwona cebulka',             'kg', 'warzywa'),
    ('PAPRYKA',           'Papryka',                      'kg', 'warzywa'),
    ('OLIWKI',            'Oliwki',                       'kg', 'warzywa'),
    ('POMIDORKI_KOKAT',   'Pomidorki koktajlowe',         'kg', 'warzywa'),
    ('POMIDORKI_SUSZE',   'Suszone pomidorki',            'kg', 'warzywa'),
    ('KUKURYDZA',         'Kukurydza',                    'kg', 'warzywa'),
    ('RUKOLA',            'Rukola',                       'kg', 'warzywa'),
    ('SZPINAK',           'Szpinak',                      'kg', 'warzywa'),
    ('CHILLI',            'Chilli (świeże)',               'kg', 'warzywa'),
    ('JALAPENO',          'Jalapeño',                     'kg', 'warzywa'),
    ('KAP_PEKINSKA',      'Kapusta pekińska',             'kg', 'warzywa'),
    ('OGUREK_KONS',       'Ogórek konserwowy',            'kg', 'warzywa'),
    ('SALATA',            'Sałata lodowa',                'kg', 'warzywa'),
    ('CUKINIA',           'Cukinia',                      'kg', 'warzywa'),
    ('BRUKOLY',           'Brokuły',                      'kg', 'warzywa'),
    ('BAZYLIA',           'Świeża bazylia',               'kg', 'warzywa'),
    ('GRUSZKA',           'Gruszka (do pizzy)',           'kg', 'warzywa'),
    ('ANANAS',            'Ananas',                       'kg', 'warzywa'),
    ('FRYTKI',            'Frytki (zamrożone)',           'kg', 'warzywa'),
    ('CHIPSY_ZIEMN',      'Chipsy ziemniaczane',          'kg', 'warzywa'),

    # PRZYPRAWY / TŁUSZCZE
    ('OREGANO',           'Oregano suszone',              'kg', 'przyprawy'),
    ('OLIWA',             'Oliwa z oliwek EVO',          'l',  'tłuszcze'),

    # MAKARONY / INNE
    ('MAKARON',           'Makaron (spaghetti)',          'kg', 'suche'),
    ('CIASTO_GYROS',      'Ciasto do gyrosa (pita)',      'kg', 'suche'),
    ('BUŁKA_PANINI',      'Bułka panini',                 'szt', 'suche'),

    # NAPOJE (dla zamówień)
    ('COCA_COLA',         'Coca-Cola 0.33l',              'szt', 'napoje'),
    ('SPRITE',            'Sprite 0.33l',                 'szt', 'napoje'),
    ('WODA_NIEGAZ',       'Woda mineralna niegazowana',   'szt', 'napoje'),
    ('WODA_GAZ',          'Woda mineralna gazowana',      'szt', 'napoje'),
    ('PIWO_BUTELKA',      'Piwo butelkowe 0.5l',          'szt', 'napoje'),

    # LODY
    ('LODY_GALKA',        'Lody (gałka)',                 'szt', 'desery'),
]

# SKU set for quick lookup
SYS_ITEMS_SKUS = {row[0] for row in SYS_ITEMS}

# Recipe ingredient mapping: keyword (lower) → (sku, qty_kg, unit)
# qty in base_unit of the ingredient (kg for solids, l for liquids, szt for pieces)
INGREDIENT_MAP = [
    # order matters — more specific first
    ('mozzarella di buffalo',  'MOZZ_BUFFALO',    0.090),
    ('buffalo mozzarella',     'MOZZ_BUFFALO',    0.090),
    ('mozzarella',             'MOZZ_FIOR',       0.100),
    ('ser',                    'MOZZ_FIOR',       0.100),
    ('ricotta',                'RICOTTA',         0.060),
    ('parmezan',               'PARMEZAN',        0.040),
    ('gorgonzola',             'GORGONZOLA',      0.040),
    ('salami picante',         'SALAMI_PICANTE',  0.060),
    ('salami',                 'SALAMI',          0.060),
    ('nduja',                  'NDUJA',           0.040),
    ('włoska kiełbasa',        'KIELB_WLOSKA',    0.060),
    ('szynka parmeńska',       'SZYNKA_PARM',     0.060),
    ('szynka',                 'SZYNKA',          0.060),
    ('kurczak',                'KURCZAK',         0.080),
    ('stek wołowy',            'STEK_WOLOWY',     0.080),
    ('boczek',                 'BOCZEK',          0.060),
    ('rwana wieprzowina',      'RWANA_WIEPRZ',    0.080),
    ('drobiowy kebab',         'KEBAB_DROBIOWY',  0.080),
    ('mięso wołowe',           'MIESO_WOLOWE',    0.080),
    ('anchois',                'ANCHOIS',         0.020),
    ('tuńczyk',                'TUNCZYK',         0.060),
    ('fileciki',               'ANCHOIS',         0.020),
    ('smażone pieczarki',      'PIECZARKI_SMAZONE',0.050),
    ('pieczarki',              'PIECZARKI',       0.050),
    ('suszone pomidorki',      'POMIDORKI_SUSZE', 0.030),
    ('pomidorki',              'POMIDORKI_KOKAT', 0.050),
    ('czerwona cebulka',       'CEBULA_CZERW',    0.030),
    ('cebulka',                'CEBULA_CZERW',    0.030),
    ('papryka',                'PAPRYKA',         0.040),
    ('oliwki',                 'OLIWKI',          0.030),
    ('kukurydza',              'KUKURYDZA',       0.040),
    ('rukola',                 'RUKOLA',          0.040),
    ('szpinak',                'SZPINAK',         0.040),
    ('chilli',                 'CHILLI',          0.010),
    ('jalapeño',               'JALAPENO',        0.010),
    ('jalapeno',               'JALAPENO',        0.010),
    ('kapusta pekińska',       'KAP_PEKINSKA',    0.040),
    ('ogórek',                 'OGUREK_KONS',     0.030),
    ('sałata',                 'SALATA',          0.040),
    ('cukinia',                'CUKINIA',         0.040),
    ('brokuły',                'BRUKOLY',         0.040),
    ('bazylia',                'BAZYLIA',         0.010),
    ('gruszka',                'GRUSZKA',         0.050),
    ('ananas',                 'ANANAS',          0.050),
    ('chipsy',                 'CHIPSY_ZIEMN',    0.060),
    ('krem balsamiczny',       'KREM_BALSAMICZNY',0.020),
    ('tabasco',                'TABASCO',         0.005),
    ('oliwa',                  'OLIWA',           0.020),
    ('sos śmietankowy',        'SOS_SMIETANKOWY', 0.060),
    ('sos',                    'SOS_POMIDOROWY',  0.080),
    ('pomidor',                'SOS_POMIDOROWY',  0.080),
]


def infer_recipe(description: str) -> list:
    """
    Parse description like 'buffalo mozzarella / pieczarki / Nduja'
    → list of {sku, quantity} dicts (in base_unit of the ingredient).
    """
    if not description:
        return []
    parts = [p.strip().lower() for p in re.split(r'[/,;]+', description)]
    lines = []
    used_skus = set()
    for part in parts:
        if not part:
            continue
        matched = False
        for keyword, sku, qty in INGREDIENT_MAP:
            if keyword in part:
                if sku not in used_skus:
                    lines.append({'sku': sku, 'qty': qty})
                    used_skus.add(sku)
                matched = True
                break
        if not matched and len(part) > 2:
            pass  # skip unrecognized ingredients silently
    # Always add basic pizza base if no sos found (for pizzas)
    return lines


# ──────────────────────────────────────────────────────────────────────────────
# MODIFIER GROUPS CONSOLIDATION
# ──────────────────────────────────────────────────────────────────────────────

# Map: original set names → consolidated group name + config
MODIFIER_GROUP_MAP = {
    'Sosy do wyboru':           ('Sos bazowy',         'SOS_BAZOWY',          'single',   True,  1, 1, 'pizzas'),
    'Dodatkowe warzywa 30 cm':  ('Dodatkowe warzywa',  'DODA_WARZYWA',        'multiple', False, 0, 0, 'pizzas'),
    'Dodatkowe warzywa 37 cm':  ('Dodatkowe warzywa',  'DODA_WARZYWA',        'multiple', False, 0, 0, 'pizzas'),
    'Dodatkowe mięsa 30 cm':    ('Dodatkowe mięsa',    'DODA_MIESA',          'multiple', False, 0, 0, 'pizzas'),
    'Dodatkowe mięsa 37 cm':    ('Dodatkowe mięsa',    'DODA_MIESA',          'multiple', False, 0, 0, 'pizzas'),
    'Dodatkowe sery 30 cm':     ('Dodatkowe sery',     'DODA_SERY',           'multiple', False, 0, 0, 'pizzas'),
    'Dodatkowe sery 37 cm':     ('Dodatkowe sery',     'DODA_SERY',           'multiple', False, 0, 0, 'pizzas'),
    'Pozostałe 30 cm':          ('Pozostałe pizza',    'POZOSTALE_PIZZA',     'multiple', False, 0, 0, 'pizzas'),
    'Pozostałe 37 cm':          ('Pozostałe pizza',    'POZOSTALE_PIZZA',     'multiple', False, 0, 0, 'pizzas'),
    'Wybierz 2 połówki 30 cm':  ('Wybierz 2 połówki', 'HALF_HALF',           'multiple', True,  2, 2, 'pizzas'),
    'Wybierz 2 połówki 37 cm':  ('Wybierz 2 połówki', 'HALF_HALF',           'multiple', True,  2, 2, 'pizzas'),
    'Sosy do pizzy':            ('Sosy dodatkowe',     'SOSY_DODATKOWE',      'multiple', False, 0, 3, 'pizzas'),
    'RODZAJ':                   ('Rodzaj ciasta',      'RODZAJ_CIASTA',       'single',   False, 0, 1, 'pizzas'),
    'USUŃ SKŁADNIK':            ('Usuń składnik',      'USUN_SKLADNIK',       'multiple', False, 0, 0, 'pizzas'),
    'Sosy do panini':           ('Sosy do panini',     'SOSY_PANINI',         'single',   True,  1, 1, 'panini'),
    'gaz/niegaz':               ('Gazowanie',          'GAZ_NIEGAZ',          'single',   False, 0, 1, 'napoje'),
    'Rodzaj SOKU':              ('Rodzaj soku',        'RODZAJ_SOKU',         'single',   True,  1, 1, 'soki'),
    'Rodzaj soku':              ('Rodzaj soku',        'RODZAJ_SOKU',         'single',   True,  1, 1, 'soki'),
    'GAŁKI LODÓW':              ('Gałki lodów',        'GALKI_LODOW',         'single',   False, 0, 3, 'desery'),
    'Ser do foccaci':           ('Ser do focacci',     'SER_FOCACCI',         'single',   False, 0, 1, 'foccacia'),
    'Dodatki do gyrosa':        ('Dodatki do gyrosa',  'DODA_GYROS',          'multiple', False, 0, 0, 'gyrosy'),
    'Dodatki':                  ('Dodatki ogólne',     'DODA_OGOLNE',         'multiple', False, 0, 0, 'general'),
    'PROMOCJA!':                ('Promocja',           'PROMOCJA',            'single',   False, 0, 1, 'general'),
    'Rodzaj calzone ':          ('Rodzaj calzone',     'RODZAJ_CALZONE',      'single',   True,  1, 1, 'calzone'),
    'Rodzaj calzone':           ('Rodzaj calzone',     'RODZAJ_CALZONE',      'single',   True,  1, 1, 'calzone'),
    'RODZAJ':                   ('Rodzaj ciasta',      'RODZAJ_CIASTA',       'single',   False, 0, 1, 'pizzas'),
    'dodatki do piwa':          ('Dodatki do piwa',    'DODA_PIWA',           'single',   False, 0, 1, 'piwa'),
}


def consolidate_modifier_groups(additions):
    """
    Returns:
        groups: dict {ascii_key: {name, ascii_key, type, required, min, max, scope,
                                  modifiers: {name: {price_30, price_37, price_base}}}}
        pricing_pairs: set of group ascii_keys that have size-differentiated pricing
    """
    groups = {}

    for mod in additions:
        orig_set = mod['set_name']

        # Normalize 'Rodzaj calzone ' (trailing space) → 'Rodzaj calzone'
        orig_set_norm = orig_set.strip()

        mapping = MODIFIER_GROUP_MAP.get(orig_set_norm) or MODIFIER_GROUP_MAP.get(orig_set)
        if not mapping:
            # Fallback: create generic group
            grp_name  = orig_set_norm
            grp_key   = make_sku(orig_set_norm)
            grp_type  = mod['type']
            grp_req   = mod['required']
            grp_min   = mod['min']
            grp_max   = mod['max']
            grp_scope = 'general'
        else:
            grp_name, grp_key, grp_type, grp_req, grp_min, grp_max, grp_scope = mapping

        if grp_key not in groups:
            groups[grp_key] = {
                'name':      grp_name,
                'ascii_key': grp_key,
                'type':      grp_type,
                'required':  grp_req,
                'min':       grp_min,
                'max':       grp_max,
                'scope':     grp_scope,
                'modifiers': {},
            }

        g = groups[grp_key]
        mod_name = mod['name']

        if mod_name not in g['modifiers']:
            g['modifiers'][mod_name] = {
                'name':       mod_name,
                'ascii_key':  make_sku(mod_name),
                'price_base': mod['price'],
                'price_30':   None,
                'price_37':   None,
            }

        # Detect size-differentiated pricing
        if '30 cm' in orig_set or '30cm' in orig_set.lower():
            g['modifiers'][mod_name]['price_30'] = mod['price']
        elif '37 cm' in orig_set or '37cm' in orig_set.lower():
            g['modifiers'][mod_name]['price_37'] = mod['price']
        else:
            g['modifiers'][mod_name]['price_base'] = mod['price']

    # For "Dodatki" group: deduplicate (appears with two price tiers in xlsx)
    # Keep only the first unique name+price combination
    if 'DODA_OGOLNE' in groups:
        seen = {}
        for mname, mdata in list(groups['DODA_OGOLNE']['modifiers'].items()):
            if mname not in seen:
                seen[mname] = mdata
        groups['DODA_OGOLNE']['modifiers'] = seen

    return groups


# ──────────────────────────────────────────────────────────────────────────────
# CATEGORY MAPPING
# ──────────────────────────────────────────────────────────────────────────────

CATEGORY_MAP = {
    'PIZZA 30 CM':  'PIZZE',
    'PIZZA 37 CM':  'PIZZE',
    'PANINI':       'PANINI',
    'MAKARONY':     'MAKARONY',
    'CALZONE':      'CALZONE',
    'FOCCACIA':     'FOCCACIA',
    'ZAPIEKANKI':   'ZAPIEKANKI',
    'GYROSY':       'GYROSY',
    'SAŁATKI':      'SAŁATKI',
    'SOSY':         'SOSY',
    'DESERY':       'DESERY',
    'DLA DZIECI':   'DLA DZIECI',
    'NAPOJE ZIMNE': 'NAPOJE',
    'PIWA':         'PIWA',
    'POZOSTAŁE':    'POZOSTAŁE',
    'NOWOŚCI !!!':  'NOWOŚCI',
    'ZIMOWE MENU':  'ZIMOWE MENU',
}

# Categories in display order
CATEGORIES_ORDER = [
    'PIZZE', 'PANINI', 'CALZONE', 'MAKARONY', 'ZAPIEKANKI',
    'GYROSY', 'FOCCACIA', 'SAŁATKI', 'SOSY', 'DESERY',
    'DLA DZIECI', 'NAPOJE', 'PIWA', 'POZOSTAŁE', 'NOWOŚCI', 'ZIMOWE MENU',
]

# VAT rates by category
VAT_MAP = {
    'PIZZE':     (8.0, 5.0),   # dine_in=8%, takeaway=5%
    'PANINI':    (8.0, 5.0),
    'MAKARONY':  (8.0, 5.0),
    'CALZONE':   (8.0, 5.0),
    'FOCCACIA':  (8.0, 5.0),
    'ZAPIEKANKI':(8.0, 5.0),
    'GYROSY':    (8.0, 5.0),
    'SAŁATKI':   (8.0, 5.0),
    'SOSY':      (8.0, 5.0),
    'DESERY':    (8.0, 5.0),
    'DLA DZIECI':(8.0, 5.0),
    'NAPOJE':    (23.0, 23.0),
    'PIWA':      (23.0, 23.0),
    'POZOSTAŁE': (8.0, 5.0),
    'NOWOŚCI':   (8.0, 5.0),
    'ZIMOWE MENU':(8.0, 5.0),
}

# Which modifier group keys apply to which categories
CATEGORY_MODIFIER_GROUPS = {
    'PIZZE':        ['SOS_BAZOWY', 'DODA_WARZYWA', 'DODA_MIESA', 'DODA_SERY',
                     'POZOSTALE_PIZZA', 'SOSY_DODATKOWE', 'HALF_HALF',
                     'RODZAJ_CIASTA', 'USUN_SKLADNIK'],
    'PANINI':       ['SOSY_PANINI'],
    'CALZONE':      ['RODZAJ_CALZONE'],
    'FOCCACIA':     ['SER_FOCACCI'],
    'GYROSY':       ['DODA_GYROS'],
    'NAPOJE':       ['GAZ_NIEGAZ', 'RODZAJ_SOKU'],
    'DESERY':       ['GALKI_LODOW'],
    'PIWA':         ['DODA_PIWA'],
}


# ──────────────────────────────────────────────────────────────────────────────
# DYNAMIC DATA — PZ, KSeF, Zamówienia
# ──────────────────────────────────────────────────────────────────────────────

# Realistic starting stock levels (qty, avco_price_grosze) for all 67 SYS_ITEMS SKU
STOCK_LEVELS = {
    'MAKA_TYP_00':      (120.000, 350),
    'MAKA_SEZAMOWA':    (15.000,  800),
    'SOS_POMIDOROWY':   (45.000,  420),
    'SOS_SMIETANKOWY':  (10.000,  620),
    'SOS_CZOSNKOWY':    (8.000,   580),
    'SOS_BBQ':          (6.000,   750),
    'SOS_MEKSYKANSKI':  (5.000,   750),
    'SOS_1000_WYSP':    (5.000,   690),
    'TABASCO':          (2.000,   2400),
    'MAJONEZ':          (8.000,   480),
    'KETCHUP':          (8.000,   380),
    'KREM_BALSAMICZNY': (3.000,   2800),
    'MOZZ_FIOR':        (30.000,  2850),
    'MOZZ_BUFFALO':     (8.000,   5200),
    'RICOTTA':          (10.000,  1800),
    'PARMEZAN':         (6.000,   6500),
    'GORGONZOLA':       (3.000,   4800),
    'FETA':             (4.000,   3200),
    'EDAMSKI':          (5.000,   2800),
    'SALAMI_PICANTE':   (10.000,  4200),
    'SALAMI':           (8.000,   3800),
    'NDUJA':            (5.000,   5500),
    'KIELB_WLOSKA':     (7.000,   3800),
    'SZYNKA_PARM':      (4.000,   8900),
    'SZYNKA':           (8.000,   2800),
    'KURCZAK':          (10.000,  1850),
    'STEK_WOLOWY':      (5.000,   4800),
    'BOCZEK':           (6.000,   2400),
    'RWANA_WIEPRZ':     (4.000,   2800),
    'KEBAB_DROBIOWY':   (5.000,   1950),
    'MIESO_WOLOWE':     (4.000,   3500),
    'ANCHOIS':          (2.000,   4500),
    'TUNCZYK':          (6.000,   2200),
    'GYROS_MIESO':      (8.000,   2200),
    'PIECZARKI':        (15.000,  750),
    'PIECZARKI_SMAZONE':(5.000,   950),
    'CEBULA_CZERW':     (10.000,  320),
    'PAPRYKA':          (8.000,   600),
    'OLIWKI':           (5.000,   1200),
    'POMIDORKI_KOKAT':  (12.000,  850),
    'POMIDORKI_SUSZE':  (4.000,   3500),
    'KUKURYDZA':        (6.000,   380),
    'RUKOLA':           (10.000,  1400),
    'SZPINAK':          (6.000,   900),
    'CHILLI':           (2.000,   1800),
    'JALAPENO':         (2.000,   2200),
    'KAP_PEKINSKA':     (4.000,   480),
    'OGUREK_KONS':      (3.000,   480),
    'SALATA':           (5.000,   800),
    'CUKINIA':          (4.000,   520),
    'BRUKOLY':          (3.000,   680),
    'BAZYLIA':          (2.000,   2500),
    'GRUSZKA':          (5.000,   650),
    'ANANAS':           (5.000,   780),
    'FRYTKI':           (8.000,   450),
    'CHIPSY_ZIEMN':     (6.000,   800),
    'OREGANO':          (3.000,   2200),
    'OLIWA':            (10.000,  1890),
    'MAKARON':          (10.000,  480),
    'CIASTO_GYROS':     (8.000,   580),
    'BUŁKA_PANINI':     (250.0,   85),
    'COCA_COLA':        (120.0,   220),
    'SPRITE':           (60.0,    220),
    'WODA_NIEGAZ':      (60.0,    110),
    'WODA_GAZ':         (60.0,    110),
    'PIWO_BUTELKA':     (180.0,   350),
    'LODY_GALKA':       (100.0,   150),
}

SUPPLIERS = [
    ('HURTOWNIA SPOŻYWCZA WARMIA Sp. z o.o.', '5252311234'),
    ('FRUKTUS — Warzywa i Owoce',              '7780012345'),
    ('DI MARCO — Włoskie Specjały',            '1230012345'),
    ('MŁYNY POLSKIE S.A.',                     '8870012345'),
    ('BROWAR REGIONALNY TYSKIE',               '6450012345'),
]

PZ_DATA = [
    {
        'doc_number': 'PZ-2026/05/FORNO-001',
        'days_ago':   14,
        'supplier':   0,  # WARMIA
        'lines': [
            ('MAKA_TYP_00',   50.000, 3.50,   8.0),
            ('SOS_POMIDOROWY',30.000, 4.20,   5.0),
            ('MOZZ_FIOR',     20.000, 28.50,  8.0),
            ('OREGANO',        2.000, 22.00,  23.0),
            ('OLIWA',         10.000, 18.90,  23.0),
            ('SALAMI_PICANTE', 8.000, 42.00,   8.0),
            ('KIELB_WLOSKA',   5.000, 38.00,   8.0),
            ('NDUJA',          4.000, 55.00,   8.0),
        ],
    },
    {
        'doc_number': 'PZ-2026/05/FORNO-002',
        'days_ago':   7,
        'supplier':   1,  # FRUKTUS
        'lines': [
            ('RUKOLA',         8.000,  14.00,  5.0),
            ('POMIDORKI_KOKAT',10.000, 8.50,   5.0),
            ('PAPRYKA',        6.000,  6.00,   5.0),
            ('PIECZARKI',      12.000, 7.50,   5.0),
            ('SZPINAK',        5.000,  9.00,   5.0),
            ('CEBULA_CZERW',   8.000,  3.20,   5.0),
            ('BAZYLIA',        2.000,  25.00,  5.0),
            ('CHILLI',         1.000,  18.00,  5.0),
            ('SALATA',         4.000,  8.00,   5.0),
        ],
    },
    {
        'doc_number': 'PZ-2026/05/FORNO-003',
        'days_ago':   3,
        'supplier':   2,  # DI MARCO
        'lines': [
            ('MOZZ_BUFFALO',   6.000,  52.00,   8.0),
            ('PARMEZAN',       4.000,  65.00,   8.0),
            ('RICOTTA',        5.000,  18.00,   8.0),
            ('SZYNKA_PARM',    3.000,  89.00,   8.0),
            ('KREM_BALSAMICZNY',2.000, 28.00,  23.0),
            ('ANCHOIS',        1.000,  45.00,   8.0),
            ('TUNCZYK',        4.000,  22.00,   8.0),
            ('POMIDORKI_SUSZE', 3.000, 35.00,   5.0),
        ],
    },
    {
        'doc_number': 'PZ-2026/05/FORNO-004',
        'days_ago':   10,
        'supplier':   3,  # MŁYNY
        'lines': [
            ('MAKA_TYP_00',    80.000, 3.50,   8.0),
            ('MAKA_SEZAMOWA',  10.000, 8.00,   8.0),
            ('BUŁKA_PANINI',   200.0,  0.85,   8.0),
        ],
    },
    {
        'doc_number': 'PZ-2026/05/FORNO-005',
        'days_ago':   1,
        'supplier':   4,  # BROWAR
        'lines': [
            ('PIWO_BUTELKA',   144.0,  3.50,  23.0),
            ('COCA_COLA',       96.0,  2.20,  23.0),
            ('SPRITE',          48.0,  2.20,  23.0),
            ('WODA_NIEGAZ',     48.0,  1.10,  23.0),
            ('WODA_GAZ',        48.0,  1.10,  23.0),
        ],
    },
]

# KSeF invoices fixture
KSEF_INVOICES = [
    {
        'invoice_number':   'FA/FORNO/2026/001',
        'supplier_nip':     '5252311234',
        'supplier_name':    'HURTOWNIA SPOŻYWCZA WARMIA Sp. z o.o.',
        'supplier_address': 'ul. Warmińska 14, 10-100 Olsztyn',
        'issue_date':        date_ago(2),
        'sale_date':         date_ago(3),
        'payment_due_date':  date_ago(-28),
        'status':           'new',
        'lines': [
            (1, 'Mąka pszenna typ 00 (25 kg worek)',        'kg', 50.0, 3.50, 8.0,  None, None, None),
            (2, 'Passata pomidorowa do pizzy',               'kg', 30.0, 4.20, 5.0,  None, None, None),
            (3, 'Mozzarella fior di latte 1 kg',            'kg', 20.0, 28.50, 8.0, None, None, None),
            (4, 'Oregano suszone premium',                  'kg',  2.0, 22.00, 23.0, None, None, None),
        ],
    },
    {
        'invoice_number':   'FA/FORNO/2026/002',
        'supplier_nip':     '7780012345',
        'supplier_name':    'FRUKTUS — Warzywa i Owoce',
        'supplier_address': 'ul. Tęczowa 3, 54-125 Wrocław',
        'issue_date':        date_ago(8),
        'sale_date':         date_ago(8),
        'payment_due_date':  date_ago(-22),
        'status':           'accepted',
        'linked_pz':        'PZ-2026/05/FORNO-002',
        'lines': [
            (1, 'Rukola świeża 500g',              'kg',  8.0, 14.00, 5.0, 'RUKOLA',         'EXACT', 99),
            (2, 'Pomidorki koktajlowe 1 kg',       'kg', 10.0,  8.50, 5.0, 'POMIDORKI_KOKAT','EXACT', 99),
            (3, 'Papryka czerwona',                'kg',  6.0,  6.00, 5.0, 'PAPRYKA',        'ALIAS', 95),
            (4, 'Pieczarki świeże 1 kg',           'kg', 12.0,  7.50, 5.0, 'PIECZARKI',      'EXACT', 99),
            (5, 'Szpinak baby liście',             'kg',  5.0,  9.00, 5.0, 'SZPINAK',        'EXACT', 99),
        ],
    },
    {
        'invoice_number':   'FA/FORNO/2026/003',
        'supplier_nip':     '1230012345',
        'supplier_name':    'DI MARCO — Włoskie Specjały',
        'supplier_address': 'ul. Włoska 8, 00-001 Warszawa',
        'issue_date':        date_ago(4),
        'sale_date':         date_ago(4),
        'payment_due_date':  date_ago(-26),
        'status':           'processing',
        'lines': [
            (1, 'Mozzarella di bufala 250g',       'kg',  6.0, 52.00, 8.0, 'MOZZ_BUFFALO',  'EXACT', 99),
            (2, 'Parmigiano Reggiano DOP 1 kg',    'kg',  4.0, 65.00, 8.0, 'PARMEZAN',      'ALIAS', 88),
            (3, 'Salsiccia Piccante 500g',         'kg',  3.0, 55.00, 8.0,  None,            'FUZZY', 62),
            (4, 'Prosciutto di Parma 1 kg',        'kg',  3.0, 89.00, 8.0,  None,             None,   None),
        ],
    },
]

# Test orders
ORDERS_DATA = [
    {
        'order_number': 'FORNO-001',
        'channel':      'delivery',
        'order_type':   'delivery',
        'status':       'accepted',
        'payment_status': 'card',
        'payment_method': 'card',
        'customer_name':  'Jan Kowalski',
        'customer_phone': '+48 512 345 678',
        'delivery_address': 'ul. Zielona 15, 10-900 Olsztyn',
        'lat':  53.7784,
        'lng':  20.4801,
        'hours_ago': 1.5,
        'delivery_fee': 800,
        'lines': [
            ('MARGHERITA_30CM',    'Margherita 30cm',      2700, 1, 8.0, None),
            ('DI_PARMA_37CM',      'Di Parma 37cm',        5400, 1, 8.0, None),
            ('COCA_COLA',          'Coca-Cola 0.33l',       700, 2, 23.0, None),
        ],
    },
    {
        'order_number': 'FORNO-002',
        'channel':      'delivery',
        'order_type':   'delivery',
        'status':       'preparing',
        'payment_status': 'online_paid',
        'payment_method': 'online',
        'customer_name':  'Anna Nowak',
        'customer_phone': '+48 601 234 567',
        'delivery_address': 'ul. Lipowa 7, 10-500 Olsztyn',
        'lat':  53.7720,
        'lng':  20.4925,
        'hours_ago': 0.75,
        'delivery_fee': 800,
        'lines': [
            ('ETNA_30CM',          'Etna 30cm',             3800, 1, 8.0, None),
            ('VINCI_30CM',         'Vinci 30cm',            3700, 1, 8.0, None),
        ],
    },
    {
        'order_number': 'FORNO-003',
        'channel':      'takeaway',
        'order_type':   'takeaway',
        'status':       'new',
        'payment_status': 'to_pay',
        'payment_method': 'cash',
        'customer_name':  'Marcin Wójcik',
        'customer_phone': '+48 789 123 456',
        'delivery_address': None,
        'lat':  None,
        'lng':  None,
        'hours_ago': 0.25,
        'delivery_fee': 0,
        'lines': [
            ('MARGHERITA_37CM',    'Margherita 37cm',       4050, 1, 8.0, None),
        ],
    },
    {
        'order_number': 'FORNO-004',
        'channel':      'pos',
        'order_type':   'dine_in',
        'status':       'preparing',
        'payment_status': 'to_pay',
        'payment_method': None,
        'customer_name':  'Stolik 4',
        'customer_phone': None,
        'delivery_address': None,
        'lat':  None,
        'lng':  None,
        'hours_ago': 0.5,
        'delivery_fee': 0,
        'lines': [
            ('MONTANARA_37CM',     'Montanara 37cm',        5550, 1, 8.0, None),
            ('STEAKY_37CM',        'Steaky 37cm',           6000, 1, 8.0, None),
            ('CAPRICCIOSA_30CM',   'Capricciosa 30cm',      3600, 1, 8.0, None),
            ('PIWO_BUTELKA',       'Piwo 0.5l',             1300, 2, 23.0, None),
        ],
    },
    {
        'order_number': 'FORNO-005',
        'channel':      'delivery',
        'order_type':   'delivery',
        'status':       'completed',
        'delivery_status': 'delivered',
        'payment_status': 'card',
        'payment_method': 'card',
        'customer_name':  'Kasia Zalewska',
        'customer_phone': '+48 698 765 432',
        'delivery_address': 'ul. Mickiewicza 33, 10-230 Olsztyn',
        'lat':  53.7801,
        'lng':  20.4756,
        'hours_ago': 168,  # 7 days ago
        'delivery_fee': 800,
        'lines': [
            ('QUATTRO_FORMAGGI_37CM', 'Quattro Formaggi 37cm', 5700, 1, 8.0, None),
            ('DI_PARMA_30CM',      'Di Parma 30cm',         3600, 1, 8.0, None),
            ('WODA_NIEGAZ',        'Woda niegazowana',       500, 2, 23.0, None),
        ],
    },
    {
        'order_number': 'FORNO-006',
        'channel':      'delivery',
        'order_type':   'delivery',
        'status':       'ready',
        'delivery_status': 'in_delivery',
        'payment_status': 'online_paid',
        'payment_method': 'online',
        'tracking_token': True,
        'customer_name':  'Piotr Nowicki',
        'customer_phone': '+48 504 321 987',
        'delivery_address': 'ul. Słoneczna 21, 10-710 Olsztyn',
        'lat':  53.7650,
        'lng':  20.5100,
        'hours_ago': 1.25,
        'delivery_fee': 800,
        'lines': [
            ('DIAVOLA_30CM',       'Diavola 30cm',          3500, 1, 8.0, None),
            ('AMERICANA_37CM',     'Americano 37cm',         5100, 1, 8.0, None),
        ],
    },
    {
        'order_number': 'FORNO-007',
        'channel':      'online',
        'order_type':   'delivery',
        'status':       'new',
        'payment_status': 'online_paid',
        'payment_method': 'online',
        'customer_name':  'Tomek Bąk',
        'customer_phone': '+48 666 555 444',
        'delivery_address': 'ul. Kościuszki 5, 10-100 Olsztyn',
        'lat':  53.7754,
        'lng':  20.4818,
        'hours_ago': 0.1,
        'delivery_fee': 800,
        'lines': [
            ('MARGHERITA_ITALIANO_30CM', 'Margherita Italiano 30cm', 3300, 1, 8.0, None),
            ('SPRITE',                   'Sprite 0.33l',              700, 1, 23.0, None),
        ],
    },
    {
        'order_number': 'FORNO-008',
        'channel':      'pos',
        'order_type':   'dine_in',
        'status':       'completed',
        'payment_status': 'card',
        'payment_method': 'card',
        'customer_name':  'Stolik 8',
        'customer_phone': None,
        'delivery_address': None,
        'lat':  None,
        'lng':  None,
        'hours_ago': 48,
        'delivery_fee': 0,
        'lines': [
            ('CARBONARA_30CM',     'Carbonara 30cm',        3600, 2, 8.0, None),
            ('VERDURA_37CM',       'Verdura 37cm',          5250, 1, 8.0, None),
            ('CAPRICCIOSA_37CM',   'Capricciosa 37cm',      5400, 1, 8.0, None),
            ('PIWO_BUTELKA',       'Piwo 0.5l',             1300, 4, 23.0, None),
            ('COCA_COLA',          'Coca-Cola 0.33l',        700, 2, 23.0, None),
        ],
    },
]


# ──────────────────────────────────────────────────────────────────────────────
# SQL BUILDER
# ──────────────────────────────────────────────────────────────────────────────

def build_sql(menu_items, additions, menu_only=False):
    lines_out = []
    W = lines_out.append

    pizza_families, panini_families, singles = extract_variant_families(menu_items)
    mod_groups = consolidate_modifier_groups(additions)

    # Track what SKUs we insert (for cleanup + validation)
    all_menu_skus = set()
    all_sys_skus  = {r[0] for r in SYS_ITEMS}

    # ── HEADER ────────────────────────────────────────────────────────────────
    now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n_pizzas  = sum(1 for v in pizza_families.values() if v['30'] or v['37'])
    n_panini  = len(panini_families)
    n_singles = len(singles)

    seed_name = 'seed_pizzaforno_menu.sql' if menu_only else 'seed_pizzaforno.sql'
    mode_label = 'MENU-ONLY (katalog jedzenia: sys_items, menu, modyfikatory, receptury)' if menu_only else 'FULL (menu + wh_stock + PZ + KSeF + zamówienia)'

    W(f"""-- =============================================================================
-- {seed_name} — SliceHub Pro
-- Wygenerowane: {now_str}
-- Źródło: _docs/menu_pizzaforno/menu (14).xlsx + additions.xlsx
-- Tryb: {mode_label}
-- Rodzin pizzy: {n_pizzas} | Panini: {n_panini} | Pojedyncze: {n_singles}
-- Składniki sys_items: {len(SYS_ITEMS)} SKU
-- Modifier groups: {len(mod_groups)}
-- =============================================================================
-- IDEMPOTENTNY — można uruchamiać wielokrotnie (cleanup na początku).
-- Zmień @tid przed uruchomieniem jeśli inny tenant.
-- Wymaga: istniejącego tenant w sh_tenant (np. utworzonego przez install_panel.php)
-- =============================================================================

SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci;
SET @tid := 2;
SET @wh  := {esc(WAREHOUSE_ID)};

""")

    # ── CLEANUP ───────────────────────────────────────────────────────────────
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("-- SEKCJA 1: CLEANUP (idempotency)")
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("SET FOREIGN_KEY_CHECKS = 0;")
    W("")

    # Orders (only in full mode — menu-only seed doesn't create orders)
    if not menu_only:
        W("DELETE FROM sh_order_payments WHERE order_id IN")
        W("  (SELECT id FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%');")
        W("DELETE FROM sh_order_audit WHERE order_id IN")
        W("  (SELECT id FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%');")
        W("DELETE FROM sh_order_lines WHERE order_id IN")
        W("  (SELECT id FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%');")
        W("DELETE FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%';")
        W("")

    # KSeF (only in full mode)
    if not menu_only:
        W("DELETE FROM sh_ksef_invoice_lines WHERE ksef_invoice_id IN")
        W("  (SELECT id FROM sh_ksef_invoices WHERE tenant_id=@tid AND invoice_number LIKE 'FA/FORNO/%');")
        W("DELETE FROM sh_ksef_invoices WHERE tenant_id=@tid AND invoice_number LIKE 'FA/FORNO/%';")
        W("")

    # WH documents (only in full mode)
    if not menu_only:
        W("DELETE FROM wh_document_lines WHERE document_id IN")
        W("  (SELECT id FROM wh_documents WHERE tenant_id=@tid AND doc_number LIKE 'PZ-2026/%/FORNO%');")
        W("DELETE FROM wh_documents WHERE tenant_id=@tid AND doc_number LIKE 'PZ-2026/%/FORNO%';")
        W("")

    # modifier_pricing for our groups
    grp_keys_quoted = ', '.join(esc(k) for k in mod_groups.keys())
    W(f"DELETE FROM sh_modifier_pricing WHERE tenant_id=@tid AND modifier_id IN")
    W(f"  (SELECT id FROM sh_modifiers WHERE group_id IN")
    W(f"   (SELECT id FROM sh_modifier_groups WHERE tenant_id=@tid AND ascii_key IN ({grp_keys_quoted})));")
    W("")

    # item_modifiers
    W("DELETE FROM sh_item_modifiers WHERE group_id IN")
    W(f"  (SELECT id FROM sh_modifier_groups WHERE tenant_id=@tid AND ascii_key IN ({grp_keys_quoted}));")
    W("")

    # modifiers
    W("DELETE FROM sh_modifiers WHERE group_id IN")
    W(f"  (SELECT id FROM sh_modifier_groups WHERE tenant_id=@tid AND ascii_key IN ({grp_keys_quoted}));")
    W("")

    # modifier groups
    W(f"DELETE FROM sh_modifier_groups WHERE tenant_id=@tid AND ascii_key IN ({grp_keys_quoted});")
    W("")

    # recipes — delete all for this tenant (seed rebuilds fully)
    W("DELETE FROM sh_recipes WHERE tenant_id=@tid AND menu_item_sku IN")
    W("  (SELECT ascii_key FROM sh_menu_items WHERE tenant_id=@tid AND")
    W("   (ascii_key LIKE '%_30CM' OR ascii_key LIKE '%_37CM' OR ascii_key LIKE '%_MALE'")
    W("   OR ascii_key LIKE '%_DUZE' OR category_id IN")
    W("   (SELECT id FROM sh_categories WHERE tenant_id=@tid AND name IN")
    W("   ('PIZZE','PANINI','CALZONE','MAKARONY','FOCCACIA','ZAPIEKANKI','GYROSY',")
    W("   'SAŁATKI','SOSY','DESERY','DLA DZIECI','NAPOJE','PIWA','POZOSTAŁE','NOWOŚCI','ZIMOWE MENU'))));")
    W("")

    # price tiers — for all our items
    W("DELETE FROM sh_price_tiers WHERE tenant_id=@tid AND target_type='ITEM' AND target_sku IN")
    W("  (SELECT ascii_key FROM sh_menu_items WHERE tenant_id=@tid AND category_id IN")
    W("   (SELECT id FROM sh_categories WHERE tenant_id=@tid AND name IN")
    W("   ('PIZZE','PANINI','CALZONE','MAKARONY','FOCCACIA','ZAPIEKANKI','GYROSY',")
    W("   'SAŁATKI','SOSY','DESERY','DLA DZIECI','NAPOJE','PIWA','POZOSTAŁE','NOWOŚCI','ZIMOWE MENU')));")
    W("")

    # menu items + categories (cascade order)
    W("DELETE FROM sh_menu_items WHERE tenant_id=@tid AND category_id IN")
    W("  (SELECT id FROM sh_categories WHERE tenant_id=@tid AND name IN")
    W("   ('PIZZE','PANINI','CALZONE','MAKARONY','FOCCACIA','ZAPIEKANKI','GYROSY',")
    W("   'SAŁATKI','SOSY','DESERY','DLA DZIECI','NAPOJE','PIWA','POZOSTAŁE','NOWOŚCI','ZIMOWE MENU'));")
    W("-- Also delete variant parents (is_variant_parent=1)")
    W("DELETE FROM sh_menu_items WHERE tenant_id=@tid AND is_variant_parent=1")
    W("  AND ascii_key NOT LIKE 'TEST%' AND ascii_key NOT LIKE 'DEMO%';")
    W("")

    W("DELETE FROM sh_categories WHERE tenant_id=@tid AND name IN")
    W("  ('PIZZE','PANINI','CALZONE','MAKARONY','FOCCACIA','ZAPIEKANKI','GYROSY',")
    W("  'SAŁATKI','SOSY','DESERY','DLA DZIECI','NAPOJE','PIWA','POZOSTAŁE','NOWOŚCI','ZIMOWE MENU');")
    W("")

    # wh_stock for our SKUs (only in full mode — menu-only doesn't manage stock)
    sys_skus_quoted = ', '.join(esc(r[0]) for r in SYS_ITEMS)
    if not menu_only:
        W(f"DELETE FROM wh_stock WHERE tenant_id=@tid AND sku IN ({sys_skus_quoted});")
        W("")
        # HR tables (must delete before sh_users — FK)
        W("DELETE FROM sh_payroll_ledger WHERE tenant_id=@tid AND employee_id IN")
        W("  (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%');")
        W("DELETE FROM sh_work_sessions WHERE tenant_id=@tid AND employee_id IN")
        W("  (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%');")
        W("DELETE FROM sh_employee_rates WHERE tenant_id=@tid AND employee_id IN")
        W("  (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%');")
        W("DELETE FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%';")
        W("")
        # sh_driver_shifts + sh_drivers
        W("DELETE FROM sh_driver_shifts WHERE driver_id IN")
        W("  (SELECT user_id FROM sh_drivers WHERE tenant_id=@tid);")
        W("DELETE FROM sh_drivers WHERE tenant_id=@tid;")
        W("")
        # sh_users
        W("DELETE FROM sh_users WHERE tenant_id=@tid AND username IN")
        W("  ('forno_owner','forno_manager','forno_waiter','forno_cook','forno_driver');")
        W("")
        # sh_tenant_settings
        W("DELETE FROM sh_tenant_settings WHERE tenant_id=@tid AND setting_key IN")
        W("  ('', 'currency', 'default_vat_dine_in', 'default_vat_takeaway', 'half_half_surcharge');")
        W("")

    # sys_items
    W(f"DELETE FROM sys_items WHERE tenant_id=@tid AND sku IN ({sys_skus_quoted});")
    W("")

    # variant scales
    W("DELETE FROM sh_variant_scale_options WHERE tenant_id=@tid AND scale_id IN")
    W("  (SELECT id FROM sh_variant_scales WHERE tenant_id=@tid AND key_ascii IN ('SCALE_PIZZA','SCALE_PANINI'));")
    W("DELETE FROM sh_variant_scales WHERE tenant_id=@tid AND key_ascii IN ('SCALE_PIZZA','SCALE_PANINI');")
    W("")

    W("SET FOREIGN_KEY_CHECKS = 1;")
    W("")

    # ── START TRANSACTION ─────────────────────────────────────────────────────
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("-- SEKCJA 2: INSERT DATA")
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("START TRANSACTION;")
    W("")

    # ── SYS_ITEMS ─────────────────────────────────────────────────────────────
    W("-- ── 2.1 sys_items (słownik surowców) ───────────────────────────────────")
    batch = []
    for sku, name, unit, cat in SYS_ITEMS:
        batch.append(f"(@tid, {esc(sku)}, {esc(name)}, {esc(unit)}, 1, 0)")
    W("INSERT INTO sys_items (tenant_id, sku, name, base_unit, is_active, is_deleted) VALUES")
    W(",\n".join(batch) + ";")
    W("")

    # ── VARIANT SCALES ────────────────────────────────────────────────────────
    W("-- ── 2.2 sh_variant_scales + sh_variant_scale_options ───────────────────")
    W("INSERT INTO sh_variant_scales (tenant_id, name, key_ascii, description, is_active)")
    W("VALUES (@tid, 'Rozmiary pizzy', 'SCALE_PIZZA', 'Skala pizzy 30cm / 37cm', 1),")
    W("       (@tid, 'Rozmiary panini', 'SCALE_PANINI', 'Skala panini małe / duże', 1);")
    W("")
    W("SET @scale_pizza_id  = (SELECT id FROM sh_variant_scales WHERE tenant_id=@tid AND key_ascii='SCALE_PIZZA');")
    W("SET @scale_panini_id = (SELECT id FROM sh_variant_scales WHERE tenant_id=@tid AND key_ascii='SCALE_PANINI');")
    W("")
    W("INSERT INTO sh_variant_scale_options (scale_id, tenant_id, name, key_ascii, display_order, multiplier, diameter_cm, is_default)")
    W("VALUES (@scale_pizza_id,  @tid, '30 cm', '30CM', 0, 1.000, 30, 1),")
    W("       (@scale_pizza_id,  @tid, '37 cm', '37CM', 1, 1.500, 37, 0),")
    W("       (@scale_panini_id, @tid, 'MAŁE',  'MALE', 0, 1.000, NULL, 1),")
    W("       (@scale_panini_id, @tid, 'DUŻE',  'DUZE', 1, 1.500, NULL, 0);")
    W("")
    W("SET @opt_30cm  = (SELECT id FROM sh_variant_scale_options WHERE scale_id=@scale_pizza_id  AND key_ascii='30CM');")
    W("SET @opt_37cm  = (SELECT id FROM sh_variant_scale_options WHERE scale_id=@scale_pizza_id  AND key_ascii='37CM');")
    W("SET @opt_male  = (SELECT id FROM sh_variant_scale_options WHERE scale_id=@scale_panini_id AND key_ascii='MALE');")
    W("SET @opt_duze  = (SELECT id FROM sh_variant_scale_options WHERE scale_id=@scale_panini_id AND key_ascii='DUZE');")
    W("")

    # ── CATEGORIES ────────────────────────────────────────────────────────────
    W("-- ── 2.3 sh_categories ──────────────────────────────────────────────────")
    cat_batch = []
    for i, cat_name in enumerate(CATEGORIES_ORDER):
        cat_batch.append(f"(@tid, {esc(cat_name)}, 1, {i}, 0)")
    W("INSERT INTO sh_categories (tenant_id, name, is_menu, display_order, is_deleted) VALUES")
    W(",\n".join(cat_batch) + ";")
    W("")

    # Set category ID variables
    for cat_name in CATEGORIES_ORDER:
        var_name = 'cat_' + make_sku(cat_name).replace('_', '').lower()[:12]
        W(f"SET @{var_name} = (SELECT id FROM sh_categories WHERE tenant_id=@tid AND name={esc(cat_name)});")
    W("")

    # Helper function to get category var name
    def catvar(name):
        return '@cat_' + make_sku(name).replace('_', '').lower()[:12]

    # ── MENU ITEMS — VARIANT PARENTS (PIZZE) ──────────────────────────────────
    W("-- ── 2.4 sh_menu_items — pizza parent items (is_variant_parent=1) ───────")

    pizza_display_order = 0
    for base_name in sorted(pizza_families.keys()):
        fam   = pizza_families[base_name]
        item30 = fam['30']
        item37 = fam['37']
        # Use description from whichever we have
        desc = (item30 or item37).get('desc', '') if (item30 or item37) else ''
        sku  = make_sku(base_name)
        vat_d, vat_t = VAT_MAP['PIZZE']
        all_menu_skus.add(sku)

        W(f"INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
        W(f"  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
        W(f"  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status)")
        W(f"VALUES (@tid, {catvar('PIZZE')}, {esc(base_name)}, {esc(sku)}, 'variant_parent', 1,")
        W(f"  1, @scale_pizza_id, NULL, NULL,")
        W(f"  {vat_d}, {vat_t}, {esc(desc)}, {pizza_display_order}, 'Live');")
        W(f"SET @p_{sku} = LAST_INSERT_ID();")
        pizza_display_order += 1

    W("")

    # ── MENU ITEMS — PIZZA CHILDREN ───────────────────────────────────────────
    W("-- ── 2.5 sh_menu_items — pizza variant children ─────────────────────────")

    child_rows_30 = []
    child_rows_37 = []

    for base_name in sorted(pizza_families.keys()):
        fam    = pizza_families[base_name]
        item30 = fam['30']
        item37 = fam['37']
        parent_var = f'@p_{make_sku(base_name)}'
        vat_d, vat_t = VAT_MAP['PIZZE']

        if item30:
            sku30  = make_sku(base_name) + '_30CM'
            price30 = int(round(item30['price'] * 100))
            desc30  = item30.get('desc', '')
            all_menu_skus.add(sku30)
            W(f"INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
            W(f"  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
            W(f"  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status)")
            W(f"VALUES (@tid, {catvar('PIZZE')}, {esc(item30['name'])}, {esc(sku30)}, 'variant', 1,")
            W(f"  0, NULL, {parent_var}, @opt_30cm,")
            W(f"  {vat_d}, {vat_t}, {esc(desc30)}, 0, 'Live');")

        if item37:
            sku37  = make_sku(base_name) + '_37CM'
            price37 = int(round(item37['price'] * 100))
            desc37  = item37.get('desc', '')
            all_menu_skus.add(sku37)
            W(f"INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
            W(f"  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
            W(f"  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status)")
            W(f"VALUES (@tid, {catvar('PIZZE')}, {esc(item37['name'])}, {esc(sku37)}, 'variant', 1,")
            W(f"  0, NULL, {parent_var}, @opt_37cm,")
            W(f"  {vat_d}, {vat_t}, {esc(desc37)}, 1, 'Live');")

    W("")

    # ── MENU ITEMS — PANINI PARENTS ───────────────────────────────────────────
    W("-- ── 2.6 sh_menu_items — panini parent items ───────────────────────────")

    panini_display_order = 0
    for base_name in sorted(panini_families.keys()):
        fam   = panini_families[base_name]
        item_male = fam['MALE']
        item_duze = fam['DUZE']
        desc = (item_male or item_duze).get('desc', '') if (item_male or item_duze) else ''
        sku  = make_sku(base_name)
        vat_d, vat_t = VAT_MAP['PANINI']
        all_menu_skus.add(sku)

        W(f"INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
        W(f"  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
        W(f"  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status)")
        W(f"VALUES (@tid, {catvar('PANINI')}, {esc(base_name)}, {esc(sku)}, 'variant_parent', 1,")
        W(f"  1, @scale_panini_id, NULL, NULL,")
        W(f"  {vat_d}, {vat_t}, {esc(desc)}, {panini_display_order}, 'Live');")
        W(f"SET @p_{sku} = LAST_INSERT_ID();")
        panini_display_order += 1

    W("")

    # ── MENU ITEMS — PANINI CHILDREN ─────────────────────────────────────────
    W("-- ── 2.7 sh_menu_items — panini variant children ───────────────────────")

    for base_name in sorted(panini_families.keys()):
        fam   = panini_families[base_name]
        item_male = fam['MALE']
        item_duze = fam['DUZE']
        parent_var = f'@p_{make_sku(base_name)}'
        vat_d, vat_t = VAT_MAP['PANINI']

        if item_male:
            sku_male = make_sku(base_name) + '_MALE'
            all_menu_skus.add(sku_male)
            W(f"INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
            W(f"  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
            W(f"  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status)")
            W(f"VALUES (@tid, {catvar('PANINI')}, {esc(item_male['name'])}, {esc(sku_male)}, 'variant', 1,")
            W(f"  0, NULL, {parent_var}, @opt_male,")
            W(f"  {vat_d}, {vat_t}, {esc(item_male.get('desc',''))}, 0, 'Live');")

        if item_duze:
            sku_duze = make_sku(base_name) + '_DUZE'
            all_menu_skus.add(sku_duze)
            W(f"INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
            W(f"  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
            W(f"  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status)")
            W(f"VALUES (@tid, {catvar('PANINI')}, {esc(item_duze['name'])}, {esc(sku_duze)}, 'variant', 1,")
            W(f"  0, NULL, {parent_var}, @opt_duze,")
            W(f"  {vat_d}, {vat_t}, {esc(item_duze.get('desc',''))}, 1, 'Live');")

    W("")

    # ── MENU ITEMS — SINGLES ──────────────────────────────────────────────────
    W("-- ── 2.8 sh_menu_items — single items (no variants) ───────────────────")

    cat_display_counters = {}
    single_batches = {}  # cat_name → list of value tuples

    for item in singles:
        cat_orig = item['category']
        cat_name = CATEGORY_MAP.get(cat_orig, cat_orig)
        sku      = make_sku(item['name'])
        vat_d, vat_t = VAT_MAP.get(cat_name, (8.0, 5.0))
        all_menu_skus.add(sku)

        if cat_name not in cat_display_counters:
            cat_display_counters[cat_name] = 0
        disp = cat_display_counters[cat_name]
        cat_display_counters[cat_name] += 1

        badge = 'new' if cat_orig == 'NOWOŚCI !!!' else 'NULL'
        badge_sql = "'new'" if badge == 'new' else 'NULL'

        # valid_from/to for ZIMOWE MENU
        valid_from = 'NULL'
        valid_to   = 'NULL'
        if cat_orig == 'ZIMOWE MENU':
            valid_from = "'2025-10-01'"
            valid_to   = "'2026-03-31'"

        if cat_name not in single_batches:
            single_batches[cat_name] = []

        single_batches[cat_name].append(
            f"(@tid, {catvar(cat_name)}, {esc(item['name'])}, {esc(sku)}, 'standard', 1,\n"
            f"  0, NULL, NULL, NULL,\n"
            f"  {vat_d}, {vat_t}, {esc(item['desc'])}, {disp}, 'Live',\n"
            f"  {badge_sql}, {valid_from}, {valid_to})"
        )

    for cat_name, rows in single_batches.items():
        W(f"-- {cat_name}")
        W("INSERT INTO sh_menu_items (tenant_id, category_id, name, ascii_key, `type`, is_active,")
        W("  is_variant_parent, variant_scale_id, parent_item_id, variant_option_id,")
        W("  vat_rate_dine_in, vat_rate_takeaway, description, display_order, publication_status,")
        W("  badge_type, valid_from, valid_to)")
        W("VALUES")
        W(",\n".join(rows) + ";")
        W("")

    # ── PRICE TIERS ───────────────────────────────────────────────────────────
    W("-- ── 2.9 sh_price_tiers ─────────────────────────────────────────────────")

    price_rows = []

    def add_prices(sku, price_zl):
        """Add POS/Takeaway/Delivery price rows for a given SKU."""
        p = round(price_zl, 2)
        price_rows.append(f"(@tid, 'ITEM', {esc(sku)}, 'POS',      {p})")
        price_rows.append(f"(@tid, 'ITEM', {esc(sku)}, 'Takeaway', {p})")
        price_rows.append(f"(@tid, 'ITEM', {esc(sku)}, 'Delivery', {p})")

    # Pizza children
    for base_name, fam in pizza_families.items():
        if fam['30']:
            add_prices(make_sku(base_name) + '_30CM', fam['30']['price'])
        if fam['37']:
            add_prices(make_sku(base_name) + '_37CM', fam['37']['price'])

    # Panini children
    for base_name, fam in panini_families.items():
        if fam['MALE']:
            add_prices(make_sku(base_name) + '_MALE', fam['MALE']['price'])
        if fam['DUZE']:
            add_prices(make_sku(base_name) + '_DUZE', fam['DUZE']['price'])

    # Singles
    for item in singles:
        add_prices(make_sku(item['name']), item['price'])

    W("INSERT INTO sh_price_tiers (tenant_id, target_type, target_sku, channel, price) VALUES")
    W(",\n".join(price_rows) + ";")
    W("")

    # ── MODIFIER GROUPS ───────────────────────────────────────────────────────
    W("-- ── 2.10 sh_modifier_groups ────────────────────────────────────────────")

    for grp_key, grp in mod_groups.items():
        typ_val    = 1 if grp['type'] == 'single' else 0
        req_val    = 1 if grp['required'] else 0
        min_sel    = grp['min']
        max_sel    = grp['max']

        W(f"INSERT INTO sh_modifier_groups (tenant_id, name, ascii_key, min_selection, max_selection,")
        W(f"  allow_multi_qty, is_active, is_deleted)")
        W(f"VALUES (@tid, {esc(grp['name'])}, {esc(grp_key)}, {min_sel}, {max_sel}, {req_val}, 1, 0);")
        W(f"SET @grp_{grp_key} = LAST_INSERT_ID();")
        W("")

    # ── MODIFIERS ─────────────────────────────────────────────────────────────
    W("-- ── 2.11 sh_modifiers ───────────────────────────────────────────────────")

    for grp_key, grp in mod_groups.items():
        if not grp['modifiers']:
            continue
        mod_rows = []
        for mod_name, mod_data in grp['modifiers'].items():
            mk  = make_sku(mod_name)
            # base price fallback
            bp = mod_data.get('price_base', 0.0) or 0.0
            mod_rows.append(
                f"(@grp_{grp_key}, {esc(mod_name)}, {esc(mk)}, 'ADD', NULL, 0, 0, {bp}, 0, 1)"
            )
        W(f"-- Modifiers for group: {grp['name']}")
        W("INSERT INTO sh_modifiers (group_id, name, ascii_key, action_type, linked_warehouse_sku,")
        W("  linked_quantity, linked_waste_percent, price, is_default, is_active) VALUES")
        W(",\n".join(mod_rows) + ";")
        W("")

    # ── MODIFIER PRICING (F-S2) ───────────────────────────────────────────────
    W("-- ── 2.12 sh_modifier_pricing (F-S2 size pricing) ───────────────────────")

    size_groups = {
        'DODA_WARZYWA':   True,
        'DODA_MIESA':     True,
        'DODA_SERY':      True,
        'POZOSTALE_PIZZA':True,
        'HALF_HALF':      True,
    }

    for grp_key, grp in mod_groups.items():
        if grp_key not in size_groups:
            continue
        mp_rows = []
        for mod_name, mod_data in grp['modifiers'].items():
            p30 = mod_data.get('price_30')
            p37 = mod_data.get('price_37')
            if p30 is None and p37 is None:
                continue
            mk = make_sku(mod_name)
            mod_id_sel = f"(SELECT id FROM sh_modifiers WHERE group_id=@grp_{grp_key} AND ascii_key={esc(mk)})"

            if p30 is not None:
                p30_gr = int(round(float(p30) * 100))
                mp_rows.append(f"(@tid, {mod_id_sel}, @opt_30cm, {p30_gr}, 0)")
            if p37 is not None:
                p37_gr = int(round(float(p37) * 100))
                mp_rows.append(f"(@tid, {mod_id_sel}, @opt_37cm, {p37_gr}, 0)")

        if mp_rows:
            W(f"-- Size pricing for group: {grp['name']}")
            W("INSERT INTO sh_modifier_pricing (tenant_id, modifier_id, variant_option_id, price_grosze, is_deleted)")
            W("VALUES")
            W(",\n".join(mp_rows) + ";")
            W("")

    # ── ITEM-MODIFIER LINKS ───────────────────────────────────────────────────
    W("-- ── 2.13 sh_item_modifiers (link items to modifier groups) ─────────────")

    # Pizza children get pizza modifier groups
    pizza_mod_groups = CATEGORY_MODIFIER_GROUPS.get('PIZZE', [])

    for base_name, fam in pizza_families.items():
        for size, item in [('30CM', fam['30']), ('37CM', fam['37'])]:
            if not item:
                continue
            sku = make_sku(base_name) + f'_{size}'
            for grp_key in pizza_mod_groups:
                if grp_key in mod_groups:
                    W(f"INSERT IGNORE INTO sh_item_modifiers (item_id, group_id)")
                    W(f"  SELECT mi.id, mg.id FROM sh_menu_items mi, sh_modifier_groups mg")
                    W(f"  WHERE mi.tenant_id=@tid AND mi.ascii_key={esc(sku)}")
                    W(f"  AND mg.tenant_id=@tid AND mg.ascii_key={esc(grp_key)};")

    W("")

    # Panini children → panini modifier groups
    for base_name, fam in panini_families.items():
        for size_key, item in [('MALE', fam['MALE']), ('DUZE', fam['DUZE'])]:
            if not item:
                continue
            sku = make_sku(base_name) + f'_{size_key}'
            for grp_key in CATEGORY_MODIFIER_GROUPS.get('PANINI', []):
                if grp_key in mod_groups:
                    W(f"INSERT IGNORE INTO sh_item_modifiers (item_id, group_id)")
                    W(f"  SELECT mi.id, mg.id FROM sh_menu_items mi, sh_modifier_groups mg")
                    W(f"  WHERE mi.tenant_id=@tid AND mi.ascii_key={esc(sku)}")
                    W(f"  AND mg.tenant_id=@tid AND mg.ascii_key={esc(grp_key)};")

    # Single items → category modifier groups
    for item in singles:
        cat_orig = item['category']
        cat_name = CATEGORY_MAP.get(cat_orig, cat_orig)
        sku      = make_sku(item['name'])
        for grp_key in CATEGORY_MODIFIER_GROUPS.get(cat_name, []):
            if grp_key in mod_groups:
                W(f"INSERT IGNORE INTO sh_item_modifiers (item_id, group_id)")
                W(f"  SELECT mi.id, mg.id FROM sh_menu_items mi, sh_modifier_groups mg")
                W(f"  WHERE mi.tenant_id=@tid AND mi.ascii_key={esc(sku)}")
                W(f"  AND mg.tenant_id=@tid AND mg.ascii_key={esc(grp_key)};")

    W("")

    # ── RECIPES ───────────────────────────────────────────────────────────────
    W("-- ── 2.14 sh_recipes (heuristic) ───────────────────────────────────────")

    recipe_count = 0
    recipe_rows = []

    # Pizza parents get the base recipe (maka + sos + mozz)
    for base_name, fam in pizza_families.items():
        item30 = fam['30']
        item37 = fam['37']
        desc   = (item30 or item37).get('desc', '') if (item30 or item37) else ''
        parent_sku = make_sku(base_name)
        lines  = infer_recipe(desc)

        # Always add mąka if no match
        if not any(l['sku'] == 'MAKA_TYP_00' for l in lines):
            lines.insert(0, {'sku': 'MAKA_TYP_00', 'qty': 0.280})

        for line in lines:
            recipe_rows.append(
                f"(@tid, {esc(parent_sku)}, {esc(line['sku'])}, {line['qty']}, 0, 0)"
            )
            recipe_count += 1

    # Single items — calzone, makarony etc. get simple recipes
    for item in singles:
        cat_name = CATEGORY_MAP.get(item['category'], item['category'])
        if cat_name not in ('MAKARONY', 'CALZONE', 'GYROSY', 'SAŁATKI'):
            continue
        sku   = make_sku(item['name'])
        desc  = item.get('desc', '')
        lines = infer_recipe(desc)
        for line in lines:
            recipe_rows.append(
                f"(@tid, {esc(sku)}, {esc(line['sku'])}, {line['qty']}, 0, 0)"
            )
            recipe_count += 1

    if recipe_rows:
        W("INSERT IGNORE INTO sh_recipes (tenant_id, menu_item_sku, warehouse_sku, quantity_base, waste_percent, is_packaging)")
        W("VALUES")
        W(",\n".join(recipe_rows) + ";")
    W("")

    if not menu_only:
        # ── WH_STOCK (initial) ───────────────────────────────────────────────────
        W("-- ── 2.15 wh_stock (stany początkowe) ──────────────────────────────────")

        stock_rows = []
        for sku, name, unit, cat in SYS_ITEMS:
            qty, avco = STOCK_LEVELS.get(sku, (10.000, 500))
            avco_dec  = round(avco / 100, 4)
            stock_rows.append(
                f"(@tid, @wh, {esc(sku)}, {qty}, {avco_dec}, {avco_dec})"
            )

        W("INSERT INTO wh_stock (tenant_id, warehouse_id, sku, quantity, current_avco_price, unit_net_cost) VALUES")
        W(",\n".join(stock_rows) + ";")
        W("")

        # ── WH_DOCUMENTS + LINES (PZ) ────────────────────────────────────────────
        W("-- ── 2.16 wh_documents + wh_document_lines (PZ receipts) ───────────────")

        for pz in PZ_DATA:
            supplier = SUPPLIERS[pz['supplier']]
            created  = days_ago(pz['days_ago'])
            W(f"INSERT INTO wh_documents (tenant_id, doc_number, type, warehouse_id, status,")
            W(f"  supplier_name, supplier_invoice, notes, created_at)")
            W(f"VALUES (@tid, {esc(pz['doc_number'])}, 'PZ', @wh, 'completed',")
            W(f"  {esc(supplier[0])}, {esc(pz['doc_number'])},")
            W(f"  {esc(f'NIP: {supplier[1]}')}, {esc(created)});")
            W(f"SET @pz_id = LAST_INSERT_ID();")
            W("")

            line_rows = []
            for sku, qty, unit_net, vat in pz['lines']:
                line_net  = round(qty * unit_net, 2)
                line_rows.append(
                    f"(@pz_id, {esc(sku)}, {qty}, {unit_net}, {line_net}, {vat}, 0, 0)"
                )

            W("INSERT INTO wh_document_lines (document_id, sku, quantity, unit_net_cost, line_net_value, vat_rate, system_qty, counted_qty)")
            W("VALUES")
            W(",\n".join(line_rows) + ";")
            W("")

            # Update wh_stock
            for sku, qty, unit_net, vat in pz['lines']:
                avco_dec = round(unit_net, 4)
                W(f"UPDATE wh_stock SET")
                W(f"  quantity = quantity + {qty},")
                W(f"  current_avco_price = (current_avco_price * quantity + {qty} * {avco_dec}) / (quantity + {qty}),")
                W(f"  unit_net_cost = {avco_dec}")
                W(f"WHERE tenant_id=@tid AND warehouse_id=@wh AND sku={esc(sku)};")
            W("")

        # ── KSeF INVOICES ─────────────────────────────────────────────────────────
        W("-- ── 2.17 sh_ksef_invoices + sh_ksef_invoice_lines ─────────────────────")

        for inv in KSEF_INVOICES:
            total_net = sum(
                int(round(l[3] * l[4] * 100))
                for l in inv['lines']
            )
            total_vat = sum(
                int(round(l[3] * l[4] * (l[5] / 100) * 100))
                for l in inv['lines']
            )
            total_gross = total_net + total_vat

            linked_wh = 'NULL'
            if inv.get('linked_pz'):
                linked_wh = f"(SELECT id FROM wh_documents WHERE tenant_id=@tid AND doc_number={esc(inv['linked_pz'])})"

            proc_at = f"'{inv['issue_date']}'" if inv['status'] == 'accepted' else 'NULL'

            W(f"INSERT INTO sh_ksef_invoices (tenant_id, supplier_nip, supplier_name, supplier_address,")
            W(f"  buyer_nip, buyer_name, invoice_number, issue_date, sale_date, payment_due_date,")
            W(f"  total_net_minor, total_vat_minor, total_gross_minor, status, linked_wh_document_id,")
            W(f"  fetched_at, processed_at)")
            W(f"VALUES (@tid, {esc(inv['supplier_nip'])}, {esc(inv['supplier_name'])}, {esc(inv['supplier_address'])},")
            W(f"  NULL, {esc('Pizzeria Forno')}, {esc(inv['invoice_number'])},")
            W(f"  {esc(inv['issue_date'])}, {esc(inv['sale_date'])}, {esc(inv['payment_due_date'])},")
            W(f"  {total_net}, {total_vat}, {total_gross},")
            W(f"  {esc(inv['status'])}, {linked_wh},")
            W(f"  {esc(days_ago(abs(int(inv['issue_date'].split('-')[2])-28) + 2))}, {proc_at});")
            W(f"SET @ksef_id = LAST_INSERT_ID();")
            W("")

            line_rows = []
            for lno, ext_name, unit, qty, unit_net, vat_rate, res_sku, match_type, confidence in inv['lines']:
                line_net  = int(round(qty * unit_net * 100))
                res_sku_s = esc(res_sku) if res_sku else 'NULL'
                mt_s      = esc(match_type) if match_type else 'NULL'
                conf_s    = str(confidence) if confidence else 'NULL'
                line_rows.append(
                    f"(@ksef_id, {lno}, {esc(ext_name)}, NULL, {esc(unit)}, {qty}, {unit_net}, {line_net}, {vat_rate},\n"
                    f"   {res_sku_s}, {mt_s}, {conf_s})"
                )

            W("INSERT INTO sh_ksef_invoice_lines (ksef_invoice_id, line_no, external_name, external_description,")
            W("  unit, qty, unit_net, line_net_minor, vat_rate, resolved_sku, match_type, match_confidence)")
            W("VALUES")
            W(",\n".join(line_rows) + ";")
            W("")

        # ── ORDERS ────────────────────────────────────────────────────────────────
        W("-- ── 2.18 sh_orders + sh_order_lines ───────────────────────────────────")
        order_ids_full = []

        for order in ORDERS_DATA:
            order_id  = gen_uuid()
            created_sql = sql_hours_ago(order['hours_ago'])
            subtotal  = sum(l[2] * l[3] for l in order['lines'])
            grand_tot = subtotal + order.get('delivery_fee', 0)
            order_type = canonical_order_type(order['order_type'])
            pay_status = canonical_payment_status(order['payment_status'], order.get('payment_method'))
            status, delivery_status = canonical_status_pair(
                order['status'], order['order_type'], order.get('delivery_status')
            )
            pm        = esc(order['payment_method']) if order.get('payment_method') else 'NULL'
            lat       = str(order['lat']) if order.get('lat') else 'NULL'
            lng       = str(order['lng']) if order.get('lng') else 'NULL'
            addr      = esc(order['delivery_address']) if order.get('delivery_address') else 'NULL'
            phone     = esc(order['customer_phone']) if order.get('customer_phone') else 'NULL'
            ds        = esc(delivery_status) if delivery_status else 'NULL'
            promised  = f"DATE_ADD({created_sql}, INTERVAL 35 MINUTE)" if order_type == 'delivery' else 'NULL'
            track_tok = (
                f"LOWER(SUBSTRING(REPLACE({esc(order_id)},'-',''), 1, 16))"
                if order.get('tracking_token') else 'NULL'
            )

            W(f"INSERT INTO sh_orders (id, tenant_id, order_number, channel, order_type, source,")
            W(f"  subtotal, delivery_fee, grand_total, status, payment_status, payment_method,")
            W(f"  delivery_status, customer_name, customer_phone, delivery_address, lat, lng,")
            W(f"  promised_time, tracking_token, created_at)")
            W(f"VALUES ({esc(order_id)}, @tid, {esc(order['order_number'])},")
            W(f"  {esc(order['channel'])}, {esc(order_type)}, {esc('seed')},")
            W(f"  {subtotal}, {order.get('delivery_fee',0)}, {grand_tot},")
            W(f"  {esc(status)}, {esc(pay_status)}, {pm},")
            W(f"  {ds}, {esc(order['customer_name'])}, {phone}, {addr}, {lat}, {lng},")
            W(f"  {promised}, {track_tok}, {created_sql});")
            W("")

            for item_sku, snap_name, unit_price, qty, vat_rate, mods in order['lines']:
                line_id   = gen_uuid()
                line_tot  = unit_price * qty
                vat_amt   = int(round(line_tot * vat_rate / (100 + vat_rate)))
                mods_json = 'NULL'
                W(f"INSERT INTO sh_order_lines (id, order_id, item_sku, snapshot_name,")
                W(f"  unit_price, quantity, line_total, vat_rate, vat_amount, modifiers_json)")
                W(f"VALUES ({esc(line_id)}, {esc(order_id)}, {esc(item_sku)}, {esc(snap_name)},")
                W(f"  {unit_price}, {qty}, {line_tot}, {vat_rate}, {vat_amt}, {mods_json});")
            W("")
            order_ids_full.append((order_id, None, status, created_sql, grand_tot, pay_status, pm))

        # ── 2.19 sh_tenant_settings + sh_users + sh_drivers (ops w full) ──────
        W("-- ── 2.19 sh_tenant_settings + sh_users + sh_drivers ──────────────────")
        # Tenant settings
        W("INSERT INTO sh_tenant_settings (tenant_id, setting_key, is_active, min_order_value,")
        W("  min_prep_time_minutes, sla_green_min, sla_yellow_min, base_prep_minutes,")
        W("  min_lead_time_minutes, setting_value)")
        W("VALUES (@tid, '', 1, 0, 30, 10, 5, 25, 30, NULL)")
        W("ON DUPLICATE KEY UPDATE is_active=1;")
        W("INSERT INTO sh_tenant_settings (tenant_id, setting_key, setting_value)")
        W("VALUES (@tid, 'currency', 'PLN'),")
        W("       (@tid, 'default_vat_dine_in', '8'),")
        W("       (@tid, 'default_vat_takeaway', '5'),")
        W("       (@tid, 'half_half_surcharge', '200')")
        W("ON DUPLICATE KEY UPDATE setting_value=VALUES(setting_value);")
        W("")
        # Users
        W("INSERT INTO sh_users (tenant_id, username, password_hash, pin_code, name,")
        W("  first_name, last_name, role, status, is_active, is_deleted)")
        W("VALUES")
        W("  (@tid, 'forno_owner',   '$2y$10$N9qo8uLOickgx2ZMRZoMy.MrqJZ4DpI1eKQZdJxKQKQKQKQKQKQK', '0000', 'Owner Forno',    'Owner',  'Forno',   'owner',   'active', 1, 0),")
        W("  (@tid, 'forno_manager', '$2y$10$N9qo8uLOickgx2ZMRZoMy.MrqJZ4DpI1eKQZdJxKQKQKQKQKQKQK', '1000', 'Manager Forno',  'Anna',   'Manager', 'manager', 'active', 1, 0),")
        W("  (@tid, 'forno_waiter',  '$2y$10$N9qo8uLOickgx2ZMRZoMy.MrqJZ4DpI1eKQZdJxKQKQKQKQKQKQK', '1111', 'Kelner Forno',   'Marek',  'Kelner',  'waiter',  'active', 1, 0),")
        W("  (@tid, 'forno_cook',    '$2y$10$N9qo8uLOickgx2ZMRZoMy.MrqJZ4DpI1eKQZdJxKQKQKQKQKQKQK', '3333', 'Kucharz Forno',  'Piotr',  'Kucharz', 'cook',    'active', 1, 0),")
        W("  (@tid, 'forno_driver',  '$2y$10$N9qo8uLOickgx2ZMRZoMy.MrqJZ4DpI1eKQZdJxKQKQKQKQKQKQK', '4444', 'Kierowca Forno', 'Tomek',  'Kierowca','driver',  'active', 1, 0)")
        W("ON DUPLICATE KEY UPDATE is_active=1, is_deleted=0;")
        W("SET @uid_owner   = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_owner');")
        W("SET @uid_manager = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_manager');")
        W("SET @uid_waiter  = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_waiter');")
        W("SET @uid_cook    = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_cook');")
        W("SET @uid_driver  = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_driver');")
        W("")
        # Drivers
        W("INSERT INTO sh_drivers (user_id, tenant_id, status) VALUES (@uid_driver, @tid, 'available')")
        W("ON DUPLICATE KEY UPDATE status='available';")
        W("INSERT INTO sh_driver_shifts (tenant_id, driver_id, initial_cash, status)")
        W("VALUES (@tid, @uid_driver, 10000, 'active')")
        W("ON DUPLICATE KEY UPDATE status='active';")
        W("")

        # ── 2.20 sh_employees + rates + work_sessions + payroll_ledger ─────────
        W("-- ── 2.20 sh_employees + rates + work_sessions + payroll_ledger ────────")
        hire_sql = "DATE_SUB(CURDATE(), INTERVAL 6 MONTH)"
        emp_data = [
            ('@uid_owner',   'EMP-FORNO-001', 'Owner Forno',    'Owner',  'Forno',   'owner',  hire_sql),
            ('@uid_manager', 'EMP-FORNO-002', 'Manager Forno',  'Anna',   'Manager', 'manager', hire_sql),
            ('@uid_waiter',  'EMP-FORNO-003', 'Kelner Forno',   'Marek',  'Kelner',  'waiter',  hire_sql),
            ('@uid_cook',    'EMP-FORNO-004', 'Kucharz Forno',  'Piotr',  'Kucharz', 'cook',    hire_sql),
            ('@uid_driver',  'EMP-FORNO-005', 'Kierowca Forno', 'Tomek',  'Kierowca','driver',  hire_sql),
        ]
        emp_rows = []
        for uid_var, code, disp, first, last, role, hire in emp_data:
            emp_rows.append(
                f"(@tid, {uid_var}, {esc(code)}, {esc(disp)}, {esc(first)}, {esc(last)},"
                f" {hire}, {esc(role)}, 'active', 'PLN')"
            )
        W("INSERT INTO sh_employees")
        W("  (tenant_id, user_id, employee_code, display_name, first_name, last_name,")
        W("   hire_date, primary_role, status, default_currency)")
        W("VALUES")
        W(",\n".join(emp_rows) + ";")
        W("SET @eid_owner   = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-001');")
        W("SET @eid_manager = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-002');")
        W("SET @eid_waiter  = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-003');")
        W("SET @eid_cook    = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-004');")
        W("SET @eid_driver  = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-005');")
        W("")
        # Employee rates
        rates_data = [
            ('@eid_manager', 2800),
            ('@eid_waiter',  2200),
            ('@eid_cook',    2500),
            ('@eid_driver',  2000),
        ]
        rate_rows = []
        for eid_var, amount in rates_data:
            rate_rows.append(
                f"(@tid, {eid_var}, 'hourly', {amount}, 'PLN', {hire_sql}, NULL, 'hiring')"
            )
        W("INSERT INTO sh_employee_rates")
        W("  (tenant_id, employee_id, rate_type, amount_minor, currency, effective_from, effective_to, reason)")
        W("VALUES")
        W(",\n".join(rate_rows) + ";")
        W("")
        # Work sessions + payroll ledger (4 months × ~20 shifts)
        import hashlib as _hl
        emp_rates = [
            ('@eid_owner',   '@uid_owner',   'EMP-FORNO-001', 0),
            ('@eid_manager', '@uid_manager', 'EMP-FORNO-002', 2800),
            ('@eid_waiter',  '@uid_waiter',  'EMP-FORNO-003', 2200),
            ('@eid_cook',    '@uid_cook',    'EMP-FORNO-004', 2500),
            ('@eid_driver',  '@uid_driver',  'EMP-FORNO-005', 2000),
        ]
        for eid_var, uid_var, emp_code, rate_minor in emp_rates:
            for month_offset in [3, 2, 1, 0]:
                base_days_ago = month_offset * 30
                shifts_this_month = 0
                for day in range(1, 29):
                    days_back = base_days_ago + day
                    if days_back < 0:
                        continue
                    if days_back % 7 in (5, 6):
                        continue
                    if shifts_this_month >= 20:
                        break
                    start_sql = f"DATE_SUB(NOW(), INTERVAL {days_back} DAY) - INTERVAL 8 HOUR"
                    end_sql   = f"DATE_SUB(NOW(), INTERVAL {days_back} DAY)"
                    total_hours = 8.0
                    uuid_seed = f"forno-{emp_code}-{month_offset}-{day}"
                    uuid_hash = _hl.md5(uuid_seed.encode()).hexdigest()
                    ws_uuid = f"{uuid_hash[:8]}-{uuid_hash[8:12]}-{uuid_hash[12:16]}-{uuid_hash[16:20]}-{uuid_hash[20:32]}"
                    W(f"INSERT INTO sh_work_sessions")
                    W(f"  (session_uuid, tenant_id, user_id, employee_id, start_time, end_time,")
                    W(f"   total_hours, clock_in_source, clock_out_source)")
                    W(f"VALUES ({esc(ws_uuid)}, @tid, {uid_var}, {eid_var},")
                    W(f"  {start_sql}, {end_sql}, {total_hours}, 'kiosk', 'kiosk')")
                    W(f"ON DUPLICATE KEY UPDATE end_time=VALUES(end_time), total_hours=VALUES(total_hours);")
                    W(f"SET @ws_id = (SELECT id FROM sh_work_sessions WHERE session_uuid={esc(ws_uuid)});")
                    if rate_minor > 0:
                        earnings_minor = int(total_hours * rate_minor)
                        period_sql = f"YEAR(DATE_SUB(NOW(), INTERVAL {days_back} DAY)), MONTH(DATE_SUB(NOW(), INTERVAL {days_back} DAY))"
                        ledger_uuid_seed = f"ledger-{emp_code}-{month_offset}-{day}"
                        ledger_hash = _hl.md5(ledger_uuid_seed.encode()).hexdigest()
                        ledger_uuid = f"{ledger_hash[:8]}-{ledger_hash[8:12]}-{ledger_hash[12:16]}-{ledger_hash[16:20]}-{ledger_hash[20:32]}"
                        W(f"INSERT INTO sh_payroll_ledger")
                        W(f"  (entry_uuid, tenant_id, employee_id, period_year, period_month,")
                        W(f"   entry_type, amount_minor, currency, hours_qty, rate_applied_minor,")
                        W(f"   ref_work_session_id, description, created_at)")
                        W(f"VALUES ({esc(ledger_uuid)}, @tid, {eid_var}, {period_sql},")
                        W(f"  'work_earnings', {earnings_minor}, 'PLN', {total_hours}, {rate_minor},")
                        W(f"  @ws_id, 'Demo shift', {end_sql})")
                        W(f"ON DUPLICATE KEY UPDATE amount_minor=VALUES(amount_minor), hours_qty=VALUES(hours_qty);")
                    W("")
                    shifts_this_month += 1
        W("")

        # ── 2.21 sh_order_audit + sh_order_payments ───────────────────────────
        W("-- ── 2.21 sh_order_audit + sh_order_payments ──────────────────────────")
        for order_id, _, _, _, _, _, _ in order_ids_full:
            W(f"INSERT INTO sh_order_audit (order_id, old_status, new_status, timestamp)")
            W(f"VALUES ({esc(order_id)}, 'new', 'accepted', NOW());")
        for order_id, _, _, _, grand_tot, pay_status, pm_var in order_ids_full:
            if pay_status in ('card', 'online_paid') and pm_var != 'NULL':
                pay_id = gen_uuid()
                W(f"INSERT INTO sh_order_payments (id, order_id, tenant_id, method, amount_grosze, tendered_grosze, transaction_id)")
                W(f"VALUES ({esc(pay_id)}, {esc(order_id)}, @tid, {pm_var}, {grand_tot}, {grand_tot},")
                W(f"  CONCAT('SEED-', UPPER(SUBSTRING(REPLACE({esc(order_id)},'-',''), 1, 12))));")
        W("")

    # ── COMMIT ────────────────────────────────────────────────────────────────
    W("COMMIT;")
    W("")

    # ── VALIDATION SELECTS ────────────────────────────────────────────────────
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("-- SEKCJA 3: WALIDACJA (quick check)")
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("SELECT 'menu_items'   AS entity, COUNT(*) AS cnt FROM sh_menu_items   WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'pizze_variants', COUNT(*) FROM sh_menu_items WHERE tenant_id=@tid AND ascii_key LIKE '%_30CM'")
    W("UNION ALL")
    W("SELECT 'modifier_groups', COUNT(*) FROM sh_modifier_groups WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'modifiers',   COUNT(*) FROM sh_modifiers WHERE group_id IN (SELECT id FROM sh_modifier_groups WHERE tenant_id=@tid)")
    W("UNION ALL")
    W("SELECT 'mod_pricing', COUNT(*) FROM sh_modifier_pricing WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'recipes',     COUNT(*) FROM sh_recipes WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'sys_items',   COUNT(*) FROM sys_items WHERE tenant_id=@tid")
    if not menu_only:
        W("UNION ALL")
        W("SELECT 'wh_stock',    COUNT(*) FROM wh_stock WHERE tenant_id=@tid")
        W("UNION ALL")
        W("SELECT 'pz_docs',     COUNT(*) FROM wh_documents WHERE tenant_id=@tid AND doc_number LIKE 'PZ-2026/%/FORNO%'")
        W("UNION ALL")
        W("SELECT 'ksef_invoices',COUNT(*) FROM sh_ksef_invoices WHERE tenant_id=@tid AND invoice_number LIKE 'FA/FORNO/%'")
        W("UNION ALL")
        W("SELECT 'orders',      COUNT(*) FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%'")
        W("UNION ALL")
        W("SELECT 'employees',   COUNT(*) FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%'")
        W("UNION ALL")
        W("SELECT 'emp_rates',   COUNT(*) FROM sh_employee_rates WHERE tenant_id=@tid AND employee_id IN (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%')")
        W("UNION ALL")
        W("SELECT 'work_sessions', COUNT(*) FROM sh_work_sessions WHERE tenant_id=@tid AND employee_id IN (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%')")
        W("UNION ALL")
        W("SELECT 'payroll_ledger', COUNT(*) FROM sh_payroll_ledger WHERE tenant_id=@tid AND employee_id IN (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%'));")
    else:
        W(";")
    W("")
    if menu_only:
        W("-- ✅ Seed Pizza Forno MENU załadowany pomyślnie (katalog jedzenia)!")
    else:
        W("-- ✅ Seed Pizza Forno FULL załadowany pomyślnie!")
    W("")

    return '\n'.join(lines_out), {
        'pizza_families':  len(pizza_families),
        'panini_families': len(panini_families),
        'singles':         len(singles),
        'sys_items':       len(SYS_ITEMS),
        'modifier_groups': len(mod_groups),
        'recipe_count':    recipe_count,
    }


# ──────────────────────────────────────────────────────────────────────────────
# OPS-ONLY SQL BUILDER
# ──────────────────────────────────────────────────────────────────────────────

def build_ops_sql():
    """Generate seed_pizzaforno_ops.sql — operational data only (no menu).

    Contains: cleanup, sh_tenant_settings, sh_users (5 PIN accounts),
    sh_drivers/sh_driver_shifts, wh_stock (67), wh_documents (5 PZ),
    sh_ksef_invoices (3), sh_orders (8), sh_order_audit (8), validation.

    Requires seed_pizzaforno_menu.sql to be imported first (food catalog).
    """
    lines_out = []
    W = lines_out.append

    now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── HEADER ────────────────────────────────────────────────────────────────
    W(f"""-- =============================================================================
-- seed_pizzaforno_ops.sql — SliceHub Pro
-- Wygenerowane: {now_str}
-- Tryb: OPS-ONLY (dane operacyjne: users, wh_stock, PZ, KSeF, zamówienia)
-- Wymaga: najpierw zaimportuj seed_pizzaforno_menu.sql (katalog jedzenia)
-- Wymaga: istniejącego tenant 2 w sh_tenant (np. utworzonego przez install_panel.php)
-- =============================================================================

SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci;
SET @tid := 2;
SET @wh  := {esc(WAREHOUSE_ID)};

""")

    # ── SEKCJA 0: CLEANUP (ops only — does NOT touch menu) ────────────────────
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("-- SEKCJA 0: CLEANUP (dane operacyjne — NIE dotyka menu)")
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("SET FOREIGN_KEY_CHECKS = 0;")
    W("")

    # Orders
    W("DELETE FROM sh_order_payments WHERE order_id IN")
    W("  (SELECT id FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%');")
    W("DELETE FROM sh_order_audit WHERE order_id IN")
    W("  (SELECT id FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%');")
    W("DELETE FROM sh_order_lines WHERE order_id IN")
    W("  (SELECT id FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%');")
    W("DELETE FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%';")
    W("")

    # KSeF
    W("DELETE FROM sh_ksef_invoice_lines WHERE ksef_invoice_id IN")
    W("  (SELECT id FROM sh_ksef_invoices WHERE tenant_id=@tid AND invoice_number LIKE 'FA/FORNO/%');")
    W("DELETE FROM sh_ksef_invoices WHERE tenant_id=@tid AND invoice_number LIKE 'FA/FORNO/%';")
    W("")

    # WH documents
    W("DELETE FROM wh_document_lines WHERE document_id IN")
    W("  (SELECT id FROM wh_documents WHERE tenant_id=@tid AND doc_number LIKE 'PZ-2026/%/FORNO%');")
    W("DELETE FROM wh_documents WHERE tenant_id=@tid AND doc_number LIKE 'PZ-2026/%/FORNO%';")
    W("")

    # wh_stock
    sys_skus_quoted = ', '.join(esc(r[0]) for r in SYS_ITEMS)
    W(f"DELETE FROM wh_stock WHERE tenant_id=@tid AND sku IN ({sys_skus_quoted});")
    W("")

    # sh_driver_shifts + sh_drivers (must delete before users)
    W("DELETE FROM sh_driver_shifts WHERE driver_id IN")
    W("  (SELECT user_id FROM sh_drivers WHERE tenant_id=@tid);")
    W("DELETE FROM sh_drivers WHERE tenant_id=@tid;")
    W("")

    # HR tables (must delete before sh_users — FK sh_employees.user_id → sh_users.id)
    W("DELETE FROM sh_payroll_ledger WHERE tenant_id=@tid AND employee_id IN")
    W("  (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%');")
    W("DELETE FROM sh_work_sessions WHERE tenant_id=@tid AND employee_id IN")
    W("  (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%');")
    W("DELETE FROM sh_employee_rates WHERE tenant_id=@tid AND employee_id IN")
    W("  (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%');")
    W("DELETE FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%';")
    W("")

    # sh_users
    W("DELETE FROM sh_users WHERE tenant_id=@tid AND username IN")
    W("  ('forno_owner','forno_manager','forno_waiter','forno_cook','forno_driver');")
    W("")

    # sh_tenant_settings (only our keys)
    W("DELETE FROM sh_tenant_settings WHERE tenant_id=@tid AND setting_key IN")
    W("  ('', 'currency', 'default_vat_dine_in', 'default_vat_takeaway', 'half_half_surcharge');")
    W("")

    W("SET FOREIGN_KEY_CHECKS = 1;")
    W("")

    # ── START TRANSACTION ─────────────────────────────────────────────────────
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("-- SEKCJA 1-8: INSERT DATA (operacyjne)")
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("START TRANSACTION;")
    W("")

    # ── SEKCJA 1: sh_tenant_settings (5 rows) ─────────────────────────────────
    W("-- ── 1 sh_tenant_settings (5 wierszy) ───────────────────────────────────")
    W("INSERT INTO sh_tenant_settings (tenant_id, setting_key, is_active, min_order_value,")
    W("  min_prep_time_minutes, sla_green_min, sla_yellow_min, base_prep_minutes,")
    W("  min_lead_time_minutes, setting_value)")
    W("VALUES (@tid, '', 1, 0, 30, 10, 5, 25, 30, NULL)")
    W("ON DUPLICATE KEY UPDATE is_active=1;")
    W("")
    W("INSERT INTO sh_tenant_settings (tenant_id, setting_key, setting_value)")
    W("VALUES (@tid, 'currency', 'PLN'),")
    W("       (@tid, 'default_vat_dine_in', '8'),")
    W("       (@tid, 'default_vat_takeaway', '5'),")
    W("       (@tid, 'half_half_surcharge', '200')")
    W("ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value);")
    W("")

    # ── SEKCJA 2: sh_users (5 kont testowych z PIN) ───────────────────────────
    W("-- ── 2 sh_users (5 kont testowych z PIN) ────────────────────────────────")
    W("INSERT INTO sh_users (tenant_id, username, password_hash, pin_code, name,")
    W("  first_name, last_name, role, status, is_active, is_deleted)")
    W("VALUES")
    W(f"  (@tid, 'forno_owner',   {esc(BCRYPT_HASH)}, '0000', 'Owner Forno',    'Owner',  'Forno',   'owner',   'active', 1, 0),")
    W(f"  (@tid, 'forno_manager', {esc(BCRYPT_HASH)}, '1000', 'Manager Forno',  'Anna',   'Manager', 'manager', 'active', 1, 0),")
    W(f"  (@tid, 'forno_waiter',  {esc(BCRYPT_HASH)}, '1111', 'Kelner Forno',   'Marek',  'Kelner',  'waiter',  'active', 1, 0),")
    W(f"  (@tid, 'forno_cook',    {esc(BCRYPT_HASH)}, '3333', 'Kucharz Forno',  'Piotr',  'Kucharz', 'cook',    'active', 1, 0),")
    W(f"  (@tid, 'forno_driver',  {esc(BCRYPT_HASH)}, '4444', 'Kierowca Forno', 'Tomek',  'Kierowca','driver',  'active', 1, 0)")
    W("ON DUPLICATE KEY UPDATE is_active=1;")
    W("")

    # Resolve user IDs
    W("SET @uid_owner   = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_owner');")
    W("SET @uid_manager = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_manager');")
    W("SET @uid_waiter  = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_waiter');")
    W("SET @uid_cook    = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_cook');")
    W("SET @uid_driver  = (SELECT id FROM sh_users WHERE tenant_id=@tid AND username='forno_driver');")
    W("")

    # sh_drivers + sh_driver_shifts
    W("INSERT INTO sh_drivers (user_id, tenant_id, status) VALUES (@uid_driver, @tid, 'available')")
    W("ON DUPLICATE KEY UPDATE status='available';")
    W("INSERT INTO sh_driver_shifts (tenant_id, driver_id, initial_cash, status)")
    W("VALUES (@tid, @uid_driver, 10000, 'active')")
    W("ON DUPLICATE KEY UPDATE status='active';")
    W("")

    # ── SEKCJA 2b: sh_employees (5 profili HR) ─────────────────────────────────
    W("-- ── 2b sh_employees (5 profili HR — kadry) ──────────────────────────────")
    # hire_date = 6 months ago (dynamic SQL, stays fresh on re-seed)
    hire_sql = "DATE_SUB(CURDATE(), INTERVAL 6 MONTH)"
    emp_data = [
        ('@uid_owner',   'EMP-FORNO-001', 'Owner Forno',    'Owner',  'Forno',   'owner',  hire_sql),
        ('@uid_manager', 'EMP-FORNO-002', 'Manager Forno',  'Anna',   'Manager', 'manager', hire_sql),
        ('@uid_waiter',  'EMP-FORNO-003', 'Kelner Forno',   'Marek',  'Kelner',  'waiter',  hire_sql),
        ('@uid_cook',    'EMP-FORNO-004', 'Kucharz Forno',  'Piotr',  'Kucharz', 'cook',    hire_sql),
        ('@uid_driver',  'EMP-FORNO-005', 'Kierowca Forno', 'Tomek',  'Kierowca','driver',  hire_sql),
    ]
    emp_rows = []
    for uid_var, code, disp, first, last, role, hire in emp_data:
        emp_rows.append(
            f"(@tid, {uid_var}, {esc(code)}, {esc(disp)}, {esc(first)}, {esc(last)},"
            f" {hire}, {esc(role)}, 'active', 'PLN')"
        )
    W("INSERT INTO sh_employees")
    W("  (tenant_id, user_id, employee_code, display_name, first_name, last_name,")
    W("   hire_date, primary_role, status, default_currency)")
    W("VALUES")
    W(",\n".join(emp_rows) + ";")
    W("")

    # Resolve employee IDs
    W("SET @eid_owner   = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-001');")
    W("SET @eid_manager = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-002');")
    W("SET @eid_waiter  = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-003');")
    W("SET @eid_cook    = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-004');")
    W("SET @eid_driver  = (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code='EMP-FORNO-005');")
    W("")

    # ── SEKCJA 2c: sh_employee_rates (stawki godzinowe) ────────────────────────
    W("-- ── 2c sh_employee_rates (stawki godzinowe — temporal) ──────────────────")
    # Rates in grosze per hour (amount_minor). effective_from = hire_date.
    # owner: no hourly rate (salaried) — skip. Others get market rates.
    rates_data = [
        ('@eid_manager', 2800),   # 28.00 PLN/h
        ('@eid_waiter',  2200),   # 22.00 PLN/h
        ('@eid_cook',    2500),   # 25.00 PLN/h
        ('@eid_driver',  2000),   # 20.00 PLN/h
    ]
    rate_rows = []
    for eid_var, amount in rates_data:
        rate_rows.append(
            f"(@tid, {eid_var}, 'hourly', {amount}, 'PLN', {hire_sql}, NULL, 'hiring')"
        )
    W("INSERT INTO sh_employee_rates")
    W("  (tenant_id, employee_id, rate_type, amount_minor, currency, effective_from, effective_to, reason)")
    W("VALUES")
    W(",\n".join(rate_rows) + ";")
    W("")

    # ── SEKCJA 2d: sh_work_sessions + sh_payroll_ledger (4 mies historii) ──────
    W("-- ── 2d sh_work_sessions + sh_payroll_ledger (4 mies × ~20 shiftów) ─────")
    W("-- Generowane dynamicznie: DATE_SUB(NOW(), INTERVAL X DAY) — świeże po re-seedzie.")
    W("")

    # Generate ~20 shifts/month × 4 months per employee (skip weekends)
    # Deterministic UUID per (employee_code, month, day) for idempotency
    import hashlib
    emp_rates = [
        ('@eid_owner',   '@uid_owner',   'EMP-FORNO-001', 0),     # no rate — skip ledger
        ('@eid_manager', '@uid_manager', 'EMP-FORNO-002', 2800),
        ('@eid_waiter',  '@uid_waiter',  'EMP-FORNO-003', 2200),
        ('@eid_cook',    '@uid_cook',    'EMP-FORNO-004', 2500),
        ('@eid_driver',  '@uid_driver',  'EMP-FORNO-005', 2000),
    ]

    ws_count = 0
    ledger_count = 0

    for eid_var, uid_var, emp_code, rate_minor in emp_rates:
        # 4 months back: 90, 60, 30, 0 days ago (approx mid-month points)
        # Generate ~20 shifts per month (Mon-Fri, ~4 weeks)
        for month_offset in [3, 2, 1, 0]:
            base_days_ago = month_offset * 30
            shifts_this_month = 0
            for day in range(1, 29):
                # Approximate weekday: day_offset from today
                days_back = base_days_ago + day
                if days_back < 0:
                    continue  # don't generate future
                # Skip weekends: use modulo of days_back
                # (approximate — not exact calendar, but good enough for demo)
                if days_back % 7 in (5, 6):
                    continue
                if shifts_this_month >= 20:
                    break

                # 8-hour shift: start 10:00, end 18:00
                start_sql = f"DATE_SUB(NOW(), INTERVAL {days_back} DAY) - INTERVAL 8 HOUR"
                end_sql   = f"DATE_SUB(NOW(), INTERVAL {days_back} DAY)"
                total_hours = 8.0

                # Deterministic UUID
                uuid_seed = f"forno-{emp_code}-{month_offset}-{day}"
                uuid_hash = hashlib.md5(uuid_seed.encode()).hexdigest()
                ws_uuid = f"{uuid_hash[:8]}-{uuid_hash[8:12]}-{uuid_hash[12:16]}-{uuid_hash[16:20]}-{uuid_hash[20:32]}"

                W(f"INSERT INTO sh_work_sessions")
                W(f"  (session_uuid, tenant_id, user_id, employee_id, start_time, end_time,")
                W(f"   total_hours, clock_in_source, clock_out_source)")
                W(f"VALUES ({esc(ws_uuid)}, @tid, {uid_var}, {eid_var},")
                W(f"  {start_sql}, {end_sql}, {total_hours}, 'kiosk', 'kiosk')")
                W(f"ON DUPLICATE KEY UPDATE end_time=VALUES(end_time), total_hours=VALUES(total_hours);")
                W(f"SET @ws_id = (SELECT id FROM sh_work_sessions WHERE session_uuid={esc(ws_uuid)});")
                W("")

                # Payroll ledger entry (only for employees with rate)
                if rate_minor > 0:
                    earnings_minor = int(total_hours * rate_minor)
                    # period_year and period_month from the shift date
                    period_sql = f"YEAR(DATE_SUB(NOW(), INTERVAL {days_back} DAY)), MONTH(DATE_SUB(NOW(), INTERVAL {days_back} DAY))"
                    ledger_uuid_seed = f"ledger-{emp_code}-{month_offset}-{day}"
                    ledger_hash = hashlib.md5(ledger_uuid_seed.encode()).hexdigest()
                    ledger_uuid = f"{ledger_hash[:8]}-{ledger_hash[8:12]}-{ledger_hash[12:16]}-{ledger_hash[16:20]}-{ledger_hash[20:32]}"

                    W(f"INSERT INTO sh_payroll_ledger")
                    W(f"  (entry_uuid, tenant_id, employee_id, period_year, period_month,")
                    W(f"   entry_type, amount_minor, currency, hours_qty, rate_applied_minor,")
                    W(f"   ref_work_session_id, description, created_at)")
                    W(f"VALUES ({esc(ledger_uuid)}, @tid, {eid_var}, {period_sql},")
                    W(f"  'work_earnings', {earnings_minor}, 'PLN', {total_hours}, {rate_minor},")
                    W(f"  @ws_id, 'Demo shift', {end_sql})")
                    W(f"ON DUPLICATE KEY UPDATE amount_minor=VALUES(amount_minor), hours_qty=VALUES(hours_qty);")
                    W("")
                    ledger_count += 1

                ws_count += 1
                shifts_this_month += 1

    W(f"-- Wygenerowano: {ws_count} sesji, {ledger_count} wpisów ledger")
    W("")

    # ── SEKCJA 3: wh_stock (67 wierszy) ───────────────────────────────────────
    W("-- ── 3 wh_stock (stany początkowe — 67 SKU) ─────────────────────────────")
    stock_rows = []
    for sku, name, unit, cat in SYS_ITEMS:
        qty, avco = STOCK_LEVELS.get(sku, (10.000, 500))
        avco_dec  = round(avco / 100, 4)
        stock_rows.append(
            f"(@tid, @wh, {esc(sku)}, {qty}, {avco_dec}, {avco_dec})"
        )
    W("INSERT INTO wh_stock (tenant_id, warehouse_id, sku, quantity, current_avco_price, unit_net_cost) VALUES")
    W(",\n".join(stock_rows) + ";")
    W("")

    # ── SEKCJA 4: wh_documents + wh_document_lines (5 PZ) ─────────────────────
    W("-- ── 4 wh_documents + wh_document_lines (5 PZ) ──────────────────────────")

    for pz in PZ_DATA:
        supplier = SUPPLIERS[pz['supplier']]
        created  = days_ago(pz['days_ago'])
        W(f"INSERT INTO wh_documents (tenant_id, doc_number, type, warehouse_id, status,")
        W(f"  supplier_name, supplier_invoice, notes, created_at)")
        W(f"VALUES (@tid, {esc(pz['doc_number'])}, 'PZ', @wh, 'completed',")
        W(f"  {esc(supplier[0])}, {esc(pz['doc_number'])},")
        W(f"  {esc(f'NIP: {supplier[1]}')}, {esc(created)});")
        W(f"SET @pz_id = LAST_INSERT_ID();")
        W("")

        line_rows = []
        for sku, qty, unit_net, vat in pz['lines']:
            line_net  = round(qty * unit_net, 2)
            line_rows.append(
                f"(@pz_id, {esc(sku)}, {qty}, {unit_net}, {line_net}, {vat}, 0, 0)"
            )

        W("INSERT INTO wh_document_lines (document_id, sku, quantity, unit_net_cost, line_net_value, vat_rate, system_qty, counted_qty)")
        W("VALUES")
        W(",\n".join(line_rows) + ";")
        W("")

        # Update wh_stock (add qty, recalc AVCO)
        for sku, qty, unit_net, vat in pz['lines']:
            avco_dec = round(unit_net, 4)
            W(f"UPDATE wh_stock SET")
            W(f"  quantity = quantity + {qty},")
            W(f"  current_avco_price = (current_avco_price * quantity + {qty} * {avco_dec}) / (quantity + {qty}),")
            W(f"  unit_net_cost = {avco_dec}")
            W(f"WHERE tenant_id=@tid AND warehouse_id=@wh AND sku={esc(sku)};")
        W("")

    # ── SEKCJA 5: sh_ksef_invoices + sh_ksef_invoice_lines (3 faktury) ────────
    W("-- ── 5 sh_ksef_invoices + sh_ksef_invoice_lines (3 faktury) ─────────────")

    for inv in KSEF_INVOICES:
        total_net = sum(
            int(round(l[3] * l[4] * 100))
            for l in inv['lines']
        )
        total_vat = sum(
            int(round(l[3] * l[4] * (l[5] / 100) * 100))
            for l in inv['lines']
        )
        total_gross = total_net + total_vat

        linked_wh = 'NULL'
        if inv.get('linked_pz'):
            linked_wh = f"(SELECT id FROM wh_documents WHERE tenant_id=@tid AND doc_number={esc(inv['linked_pz'])})"

        proc_at = f"'{inv['issue_date']}'" if inv['status'] == 'accepted' else 'NULL'

        W(f"INSERT INTO sh_ksef_invoices (tenant_id, supplier_nip, supplier_name, supplier_address,")
        W(f"  buyer_nip, buyer_name, invoice_number, issue_date, sale_date, payment_due_date,")
        W(f"  total_net_minor, total_vat_minor, total_gross_minor, status, linked_wh_document_id,")
        W(f"  fetched_at, processed_at)")
        W(f"VALUES (@tid, {esc(inv['supplier_nip'])}, {esc(inv['supplier_name'])}, {esc(inv['supplier_address'])},")
        W(f"  NULL, {esc('Pizzeria Forno')}, {esc(inv['invoice_number'])},")
        W(f"  {esc(inv['issue_date'])}, {esc(inv['sale_date'])}, {esc(inv['payment_due_date'])},")
        W(f"  {total_net}, {total_vat}, {total_gross},")
        W(f"  {esc(inv['status'])}, {linked_wh},")
        W(f"  {esc(days_ago(abs(int(inv['issue_date'].split('-')[2])-28) + 2))}, {proc_at});")
        W(f"SET @ksef_id = LAST_INSERT_ID();")
        W("")

        line_rows = []
        for lno, ext_name, unit, qty, unit_net, vat_rate, res_sku, match_type, confidence in inv['lines']:
            line_net  = int(round(qty * unit_net * 100))
            res_sku_s = esc(res_sku) if res_sku else 'NULL'
            mt_s      = esc(match_type) if match_type else 'NULL'
            conf_s    = str(confidence) if confidence else 'NULL'
            line_rows.append(
                f"(@ksef_id, {lno}, {esc(ext_name)}, NULL, {esc(unit)}, {qty}, {unit_net}, {line_net}, {vat_rate},\n"
                f"   {res_sku_s}, {mt_s}, {conf_s})"
            )

        W("INSERT INTO sh_ksef_invoice_lines (ksef_invoice_id, line_no, external_name, external_description,")
        W("  unit, qty, unit_net, line_net_minor, vat_rate, resolved_sku, match_type, match_confidence)")
        W("VALUES")
        W(",\n".join(line_rows) + ";")
        W("")

    # ── SEKCJA 6: sh_orders + sh_order_lines (8 zamówień) ─────────────────────
    W("-- ── 6 sh_orders + sh_order_lines (8 zamówień) ──────────────────────────")

    order_ids = []  # (order_id, user_id_sql, status, created_sql) for audit
    for order in ORDERS_DATA:
        order_id  = gen_uuid()
        created_sql = sql_hours_ago(order['hours_ago'])
        subtotal  = sum(l[2] * l[3] for l in order['lines'])
        grand_tot = subtotal + order.get('delivery_fee', 0)
        order_type = canonical_order_type(order['order_type'])
        pay_status = canonical_payment_status(order['payment_status'], order.get('payment_method'))
        status, delivery_status = canonical_status_pair(
            order['status'], order['order_type'], order.get('delivery_status')
        )
        pm        = esc(order['payment_method']) if order.get('payment_method') else 'NULL'
        lat       = str(order['lat']) if order.get('lat') else 'NULL'
        lng       = str(order['lng']) if order.get('lng') else 'NULL'
        addr      = esc(order['delivery_address']) if order.get('delivery_address') else 'NULL'
        phone     = esc(order['customer_phone']) if order.get('customer_phone') else 'NULL'
        ds        = esc(delivery_status) if delivery_status else 'NULL'
        promised  = f"DATE_ADD({created_sql}, INTERVAL 35 MINUTE)" if order_type == 'delivery' else 'NULL'
        track_tok = (
            f"LOWER(SUBSTRING(REPLACE({esc(order_id)},'-',''), 1, 16))"
            if order.get('tracking_token') else 'NULL'
        )
        # Assign user_id: waiter for dine_in, driver for delivery, NULL otherwise
        if order_type == 'dine_in':
            uid_sql = '@uid_waiter'
        elif order_type == 'delivery':
            uid_sql = '@uid_driver'
        else:
            uid_sql = 'NULL'

        W(f"INSERT INTO sh_orders (id, tenant_id, order_number, channel, order_type, source,")
        W(f"  subtotal, delivery_fee, grand_total, status, payment_status, payment_method,")
        W(f"  delivery_status, customer_name, customer_phone, delivery_address, lat, lng,")
        W(f"  promised_time, tracking_token, created_at, user_id)")
        W(f"VALUES ({esc(order_id)}, @tid, {esc(order['order_number'])},")
        W(f"  {esc(order['channel'])}, {esc(order_type)}, {esc('seed')},")
        W(f"  {subtotal}, {order.get('delivery_fee',0)}, {grand_tot},")
        W(f"  {esc(status)}, {esc(pay_status)}, {pm},")
        W(f"  {ds}, {esc(order['customer_name'])}, {phone}, {addr}, {lat}, {lng},")
        W(f"  {promised}, {track_tok}, {created_sql}, {uid_sql});")
        W("")

        for item_sku, snap_name, unit_price, qty, vat_rate, mods in order['lines']:
            line_id   = gen_uuid()
            line_tot  = unit_price * qty
            vat_amt   = int(round(line_tot * vat_rate / (100 + vat_rate)))
            mods_json = 'NULL'
            W(f"INSERT INTO sh_order_lines (id, order_id, item_sku, snapshot_name,")
            W(f"  unit_price, quantity, line_total, vat_rate, vat_amount, modifiers_json)")
            W(f"VALUES ({esc(line_id)}, {esc(order_id)}, {esc(item_sku)}, {esc(snap_name)},")
            W(f"  {unit_price}, {qty}, {line_tot}, {vat_rate}, {vat_amt}, {mods_json});")
        W("")

        order_ids.append((order_id, uid_sql, status, created_sql, grand_tot, pay_status, pm))

    # ── SEKCJA 7: sh_order_audit (8 wierszy) ──────────────────────────────────
    W("-- ── 7 sh_order_audit (8 wierszy) ───────────────────────────────────────")

    for order_id, uid_sql, status, created_sql, _, _, _ in order_ids:
        W(f"INSERT INTO sh_order_audit (order_id, user_id, old_status, new_status, timestamp)")
        W(f"VALUES ({esc(order_id)}, {uid_sql}, 'new', {esc(status)}, {created_sql});")
    W("")

    # ── SEKCJA 7b: sh_order_payments (płatności za zamówienia) ────────────────
    W("-- ── 7b sh_order_payments (płatności) ───────────────────────────────────")
    for order_id, _, _, created_sql, grand_tot, pay_status, pm_var in order_ids:
        # Only create payment records for paid orders (card, online_paid)
        # Skip 'to_pay' — payment pending, no record yet
        if pay_status in ('card', 'online_paid') and pm_var != 'NULL':
            pay_id = gen_uuid()
            method = pm_var.strip("'\"")
            W(f"INSERT INTO sh_order_payments (id, order_id, tenant_id, method, amount_grosze, tendered_grosze, transaction_id)")
            W(f"VALUES ({esc(pay_id)}, {esc(order_id)}, @tid, {pm_var}, {grand_tot}, {grand_tot},")
            W(f"  CONCAT('SEED-', UPPER(SUBSTRING(REPLACE({esc(order_id)},'-',''), 1, 12))));")
    W("")

    # ── COMMIT ────────────────────────────────────────────────────────────────
    W("COMMIT;")
    W("")

    # ── SEKCJA 8: WALIDACJA ───────────────────────────────────────────────────
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("-- SEKCJA 8: WALIDACJA (quick check)")
    W("-- ═══════════════════════════════════════════════════════════════════════")
    W("SELECT 'tenant_settings' AS entity, COUNT(*) AS cnt FROM sh_tenant_settings WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'users', COUNT(*) FROM sh_users WHERE tenant_id=@tid AND username LIKE 'forno_%'")
    W("UNION ALL")
    W("SELECT 'wh_stock', COUNT(*) FROM wh_stock WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'pz_docs', COUNT(*) FROM wh_documents WHERE tenant_id=@tid AND doc_number LIKE 'PZ-2026/%/FORNO%'")
    W("UNION ALL")
    W("SELECT 'ksef_invoices', COUNT(*) FROM sh_ksef_invoices WHERE tenant_id=@tid AND invoice_number LIKE 'FA/FORNO/%'")
    W("UNION ALL")
    W("SELECT 'orders', COUNT(*) FROM sh_orders WHERE tenant_id=@tid AND order_number LIKE 'FORNO-%'")
    W("UNION ALL")
    W("SELECT 'order_payments', COUNT(*) FROM sh_order_payments WHERE tenant_id=@tid")
    W("UNION ALL")
    W("SELECT 'employees', COUNT(*) FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%'")
    W("UNION ALL")
    W("SELECT 'employee_rates', COUNT(*) FROM sh_employee_rates WHERE tenant_id=@tid AND employee_id IN (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%')")
    W("UNION ALL")
    W("SELECT 'work_sessions', COUNT(*) FROM sh_work_sessions WHERE tenant_id=@tid AND employee_id IN (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%')")
    W("UNION ALL")
    W("SELECT 'payroll_ledger', COUNT(*) FROM sh_payroll_ledger WHERE tenant_id=@tid AND employee_id IN (SELECT id FROM sh_employees WHERE tenant_id=@tid AND employee_code LIKE 'EMP-FORNO-%'));")
    W("")
    W("-- ✅ Seed Pizza Forno OPS załadowany pomyślnie (dane operacyjne)!")
    W("")

    return '\n'.join(lines_out)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description='SliceHub Seed Builder — Pizza Forno')
    parser.add_argument('--menu-only', action='store_true',
                        help='Generuj tylko katalog jedzenia (sys_items, menu, modyfikatory, receptury). '
                             'Bez wh_stock, PZ, KSeF, zamówień. Output: seed_pizzaforno_menu.sql')
    parser.add_argument('--ops-only', action='store_true',
                        help='Generuj tylko dane operacyjne (users, wh_stock, PZ, KSeF, zamówienia). '
                             'Bez menu. Output: seed_pizzaforno_ops.sql')
    args = parser.parse_args()

    if args.menu_only and args.ops_only:
        parser.error('--menu-only i --ops-only są wzajemnie wykluczające (wybierz jeden tryb).')

    menu_only = args.menu_only
    ops_only  = args.ops_only

    if menu_only:
        output_file = MENU_OUTPUT_FILE
    elif ops_only:
        output_file = OPS_OUTPUT_FILE
    else:
        output_file = OUTPUT_FILE

    mode_label = 'MENU-ONLY' if menu_only else ('OPS-ONLY' if ops_only else 'FULL')
    print(f"🍕 SliceHub Seed Builder — Pizza Forno [{mode_label}]", file=sys.stderr)
    print(f"   Output:         {output_file}", file=sys.stderr)
    print("", file=sys.stderr)

    # OPS-ONLY mode doesn't need xlsx parsing — uses module-level constants only
    if ops_only:
        print("  [1/2] Building SQL (ops-only, no xlsx parse needed)...", file=sys.stderr)
        sql = build_ops_sql()
        print(f"  [2/2] Writing {output_file}...", file=sys.stderr)
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(sql)
        size_kb = os.path.getsize(output_file) / 1024
        print(f"        → {size_kb:.1f} KB written", file=sys.stderr)
        print("", file=sys.stderr)
        print(f"✅ Done! Import {output_file} into phpMyAdmin.", file=sys.stderr)
        print(f"   ⚠️  Najpierw zaimportuj seed_pizzaforno_menu.sql (katalog jedzenia)!", file=sys.stderr)
        return

    print(f"   Menu file:      {MENU_FILE}", file=sys.stderr)
    print(f"   Additions file: {ADD_FILE}", file=sys.stderr)
    print("", file=sys.stderr)

    print("  [1/4] Parsing menu xlsx...", file=sys.stderr)
    menu_items = parse_menu_xlsx()
    print(f"        → {len(menu_items)} items loaded", file=sys.stderr)

    print("  [2/4] Parsing additions xlsx...", file=sys.stderr)
    additions = parse_additions_xlsx()
    print(f"        → {len(additions)} modifiers loaded", file=sys.stderr)

    print("  [3/4] Building SQL...", file=sys.stderr)
    sql, stats = build_sql(menu_items, additions, menu_only=menu_only)
    print(f"        → Pizza families:  {stats['pizza_families']}", file=sys.stderr)
    print(f"          Panini families: {stats['panini_families']}", file=sys.stderr)
    print(f"          Single items:    {stats['singles']}", file=sys.stderr)
    print(f"          sys_items SKU:   {stats['sys_items']}", file=sys.stderr)
    print(f"          Modifier groups: {stats['modifier_groups']}", file=sys.stderr)
    print(f"          Recipe lines:    {stats['recipe_count']}", file=sys.stderr)

    print(f"  [4/4] Writing {output_file}...", file=sys.stderr)
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(sql)

    size_kb = os.path.getsize(output_file) / 1024
    print(f"        → {size_kb:.1f} KB written", file=sys.stderr)
    print("", file=sys.stderr)
    print(f"✅ Done! Import {output_file} into phpMyAdmin.", file=sys.stderr)


if __name__ == '__main__':
    main()
